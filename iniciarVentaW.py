import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk
import pyodbc
import re
import copy
import os
import urllib.request
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ------------------------------------------------------------
# 1. CONFIGURACIÓN DB
# ------------------------------------------------------------
DB_IP = "192.168.0.100"
DB_USER = "debo"
DB_PASS = "debo"
DB_NAME = "DEBO"

#----------------------------------------Globales--------------

PLANILLA_DESDE_SQL = None
PLANILLA_HASTA_SQL = None
CONSULTAR_SQL_HANDLER = None
EJECUTAR_SQL_HANDLER = None
TOTAL_CAJA_CONJUNTA_FN = None
DIFERENCIAS_VENDEDORES = {}
TOTAL_CAJA_CONJUNTA_VALOR = 0.0
RESPONSABLE_GNC_ACTUAL = None



sf_global = None

UNDO_STACK = []
DATOS_PRODUCTOS_FACTURADOS = None
DATOS_DETALLE_REMITOS = {}
DATOS_DETALLE_FACTURACION = {}
DATOS_RENDICIONES = {}
DF_RENDICIONES_CACHE = None
COLOR_MAS = "#3498db"      # azul
COLOR_EDIT = "#f1c40f"    # amarillo
CAJAS_DATA = {}
DESCUENTOS_QR_POR_VENDEDOR = {}
# =========================
# GNC - ESTRUCTURAS BASE
# =========================

# =========================
# GNC - SQL CACHE
# =========================
DF_GNC_SQL_CACHE = None

PRECIO_GNC_DEFAULT = 669
# =========================
# GNC - ESTADO GLOBAL
# =========================
UNDO_STACK = []
btn_deshacer = None
TRANSACCION_QR_BUSCADA = None
TURNO_SELECCIONADO = None

ANOTACIONES_TMP = []  # solo memoria, se pierde al cerrar
LABELS_ANOTACIONES = {}           # labels visuales por vendedor
TREE_ANOTACIONES = None
QR_REASIGNADOS = set()
# -------------------------
# ESTADOS QR (GLOBAL)
# -------------------------
ESTADO_QR_PENDIENTE = "PENDIENTE"
ESTADO_QR_IMPACTADO = "IMPACTO_QR"
ESTADO_QR_SIN_COMPROBANTE = "SIN_COMPROBANTE"
ESTADO_QR_MANUAL = "ASIGNADO_MANUAL"
ESTADO_QR_ASIGNADO_MANUAL = "ASIGNADO_MANUAL"




GNC_GENERAL_TOTAL = 0.0
GNC_COBERTURAS = []            # lista de CoberturaGNC
GNC_TOTAL_COBERTURAS = 0.0
GNC_PARA_CAJA = 0.0
HEADER_CREADO = False
frame_header = None
DATOS_DETALLE_QR = {}
CALCULAR_HANDLER = None
VENTANA_ANOTACIONES = None
TREE_ANOTACIONES = None

HISTORIAL_CAJAS = []

GNC_EXTRA_POR_RESPONSABLE = {} # responsable -> monto
RESPONSABLES_GNC_VALIDOS = [
    "RICARDO RAMON",
    "KARINA ADRIANA",
    "HEBER AGUSTIN",
    "RAMON ELECTO",
    "JAVIER ANTONIO",
    "MIGUEL ALEJANDRO",
    "MATIAS CASAS",
    "PONCE SANTIAGO",
    "ESTEBAN ORTIZ",
    "MARLENE GONZALEZ",
    "ARCE RUBEN EMANUEL",
    "SABRINA RUIZ",
    "ARGUELLO MARCOS",
    "ELIAS GARCIA",
    "LUCAS ROMERO",
    "NICOLAS SUAREZ",
    "ROMAN VELAZQUEZ",
    "BRIAN VEGA"
]

ANOTACIONES = {
    # nro_transaccion: {
    #     "tipo": "INFORMAR" | "DESCONTAR",
    #     "vendedor": "NOMBRE",
    #     "monto": float | None,
    #     "detalle": str,
    #     "estado": "PENDIENTE" | "RESUELTA",
    #     "qr_data": dict | None   # fila QR cuando se resuelve
    # }
}
#----------------------------------------------------------------------------------------------------------------------



class AforadorGNC:
    def __init__(self, inicial=0.0, final=0.0, precio=PRECIO_GNC_DEFAULT):
        self.inicial = float(inicial or 0)
        self.final = float(final or 0)
        self.precio = float(precio or PRECIO_GNC_DEFAULT)

    def consumo(self):
        return max(self.final - self.inicial, 0)

    def total(self):
        return self.consumo() * self.precio


class CoberturaGNC:
    """
    Usado para:
    - Cubrió GNC
    - Baños
    (misma lógica)
    """
    def __init__(self, nombre, responsable, aforadores):
        self.nombre = nombre
        self.responsable = responsable
        self.aforadores = aforadores  # lista de AforadorGNC

    def total(self):
        return sum(a.total() for a in self.aforadores)


def calcular_gnc_general(aforadores):
    """
    GNC total del turno (NO afectado por coberturas)
    """
    return sum(a.total() for a in aforadores)

def obtener_conexion_sql():
    """Intenta conectar a la base de datos SQL Server y maneja errores."""
    try:
        conn_str = (
            f"DRIVER={{SQL Server}};SERVER={DB_IP};DATABASE={DB_NAME};"
            f"UID={DB_USER};PWD={DB_PASS};Connection Timeout=10;"
        )
        return pyodbc.connect(conn_str)
    except Exception as e:
        messagebox.showerror("Error de Red", f"No se pudo conectar a la DB:\n{e}")
        return None


# ------------------------------------------------------------
# 2. HERRAMIENTAS
# ------------------------------------------------------------
def actualizar_diferencias_ui():
        for v, diff in DIFERENCIAS_VENDEDORES.items():
            if v in widgets:
                widgets[v]["diff"].config(text=f"${diff:,.2f}")
def recalcular_totales_caja(cid):
    total_qr = 0.0
    total_prod = 0.0
    total_per = 0.0
    total_tarj = 0.0
    total_ef = 0.0
    total_anot = 0.0
    total_base = 0.0

    for v, w in widgets.items():
        if w["caja_id"] != cid:
            continue

        total_qr += parse_moneda_robusto(w.get("qr_db", 0))
        total_prod += parse_moneda_robusto(w["prod"].get())
        total_per += parse_moneda_robusto(w["per"].get())
        total_tarj += parse_moneda_robusto(w["tarj"].get())
        total_ef += parse_moneda_robusto(w["ef"].get())
        total_anot += total_anotaciones_por_vendedor(v)
        total_base += w["base"]

    return {
        "qr": total_qr,
        "prod": total_prod,
        "per": total_per,
        "tarj": total_tarj,
        "ef": total_ef,
        "anot": total_anot,
        "base": total_base
    }
def reflejar_totales_en_ui(vendedor_principal):
    w = widgets[vendedor_principal]
    cid = w["caja_id"]

    tot = recalcular_totales_caja(cid)

    def set_entry(e, val):
        e.config(state="normal")
        e.delete(0, tk.END)
        e.insert(0, f"{val:.2f}")
        e.config(state="disabled")

    set_entry(w["qr"], tot["qr"])
    set_entry(w["prod"], tot["prod"])
    set_entry(w["per"], tot["per"])
    set_entry(w["tarj"], tot["tarj"])
    set_entry(w["ef"], tot["ef"])

    # Diferencia REAL
    salida = tot["qr"] + tot["tarj"] + tot["ef"]
    entrada = tot["base"] + tot["prod"] + tot["per"] - tot["anot"]

    diff = salida - entrada
    w["diff"].config(
        text=formatear_arg(diff),
        fg="#27ae60" if diff >= 0 else "#e74c3c"
    )


def desactivar_fila(v):
    w = widgets[v]
    w["fusionado"] = True

    w["row"].configure(bg="#dcdcdc")
    for c in w["row"].winfo_children():
        try:
            c.configure(state="disabled")
        except:
            pass

def recalcular_por_cajas():
    cajas = {}

    # Agrupar vendedores por caja_id
    for v, w in widgets.items():
        cid = tuple(sorted(w["caja_id"]))
        cajas.setdefault(cid, []).append((v, w))

    # Calcular por cada caja
    for cid, lista in cajas.items():

        # =========================
        # ENTRADAS
        # =========================
        entrada = 0.0
        for v, w in lista:
            entrada += (
                w["base"]
                + parse_moneda_robusto(w["prod"].get())
                + parse_moneda_robusto(w["per"].get())
            )

        # =========================
        # SALIDAS (LO QUE VE LA UI)
        # =========================
        salida = 0.0
        for v, w in lista:
            salida += (
                parse_moneda_robusto(w["qr"].get()) +
                parse_moneda_robusto(w["tarj"].get()) +
                parse_moneda_robusto(w["ef"].get())
            )

        # =========================
        # AJUSTES
        # =========================
        ajustes = 0.0
        for v, _ in lista:
            ajustes += total_anotaciones_por_vendedor(v)
            ajustes += DESCUENTOS_QR_POR_VENDEDOR.get(v, 0.0)

        diferencia = salida - (entrada - ajustes)

        # =========================
        # MOSTRAR SOLO UNA VEZ
        # =========================
        fila_principal = None
        for v, w in lista:
            if not w.get("fusionado"):
                fila_principal = w
                break

        # limpiar diffs secundarios
        for _, w in lista:
            w["diff"].config(text="")

        if fila_principal:
            fila_principal["diff"].config(
                text=formatear_arg(diferencia),
                fg="#c0392b" if diferencia < 0 else "#27ae60"
            )

def texto_anotaciones_por_vendedor(vendedor):
    textos = []

    for a in ANOTACIONES_TMP:
        if a.get("vendedor") == vendedor:
            desc = a.get("descripcion", "").strip()
            if desc:
                textos.append(f"• {desc}")

    return "\n".join(textos)
def recalcular_desde_datos(vendedor):
    w = widgets[vendedor]

    # QR
    total_qr = 0.0
    if vendedor in DATOS_DETALLE_QR:
        total_qr = DATOS_DETALLE_QR[vendedor]["QR_FINAL"].sum()

    w["qr"].config(state="normal")
    w["qr"].delete(0, tk.END)
    w["qr"].insert(0, formatear_moneda_ui(total_qr))
    w["qr"].config(state="readonly")

    # PRODUCTOS
    total_prod = 0.0
    if vendedor in DATOS_DETALLE_FACTURACION:
        total_prod = DATOS_DETALLE_FACTURACION[vendedor]["Total_Neto"].sum()

    w["prod"].config(state="normal")
    w["prod"].delete(0, tk.END)
    w["prod"].insert(0, formatear_moneda_ui(total_prod))
    w["prod"].config(state="readonly")

    # EFECTIVO
    actualizar_visual_efectivo(vendedor)

    # DIFERENCIA FINAL
    recalcular_diferencia(vendedor)

def unir_cajas_simple(vend_a, vend_b):
    w_a = widgets[vend_a]
    w_b = widgets[vend_b]

    def val(entry):
        return parse_moneda_robusto(entry.get())

    def set_entry(entry, valor):
        entry.config(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, formatear_moneda_ui(valor))
        entry.config(state="readonly")

    # =========================
    # 1️⃣ SUMAS (UI → UI)
    # =========================
    suma_qr   = val(w_a["qr"])   + val(w_b["qr"])
    suma_prod = val(w_a["prod"]) + val(w_b["prod"])
    suma_per  = val(w_a["per"])  + val(w_b["per"])
    suma_tarj = val(w_a["tarj"]) + val(w_b["tarj"])
    suma_ef   = val(w_a["ef"])   + val(w_b["ef"])

    # Venta Excel (base)
    w_a["base"] = w_a["base"] + w_b["base"]

    # =========================
    # 2️⃣ IMPACTAR EN UI (A)
    # =========================
    set_entry(w_a["qr"],   suma_qr)
    set_entry(w_a["prod"], suma_prod)
    set_entry(w_a["per"],  suma_per)
    set_entry(w_a["tarj"], suma_tarj)
    set_entry(w_a["ef"],   suma_ef)

    w_a["lbl_base"].config(
        text=formatear_moneda_ui(w_a["base"]),
        fg="#16a085"
    )

    # =========================
    # 3️⃣ MARCAR B COMO FUSIONADO
    # =========================
    w_b["fusionado"] = True
    w_b["caja_id"] = w_a["caja_id"]

    w_b["row"].configure(bg="#dcdcdc")
    for c in w_b["row"].winfo_children():
        try:
            c.configure(state="disabled")
        except:
            pass

    # =========================
    # 4️⃣ RECALCULAR DIFERENCIA
    # =========================
    recalcular_diferencia(vend_a)


def recalcular_diferencia(vendedor):
    w = widgets.get(vendedor)
    if not w:
        return

    base = w["base"]

    total_ingresado = (
        parse_moneda_robusto(w["qr"].get()) +
        parse_moneda_robusto(w["prod"].get()) +
        parse_moneda_robusto(w["per"].get()) +
        parse_moneda_robusto(w["tarj"].get()) +
        parse_moneda_robusto(w["ef"].get())
    )

    diff = base - total_ingresado
    w["diff"].config(text=formatear_arg(diff))

def total_qr_caja(cid):
    df = CAJAS_DATA[cid]["qr"]
    return 0 if df is None else df["QR_FINAL"].sum()

def abrir_detalle_qr(vendedor, root):
    w = widgets.get(vendedor)

    # si la caja está unida → usamos la nueva
    if w and isinstance(w.get("caja_id"), frozenset) and len(w["caja_id"]) > 1:
        return ver_detalle_qr_caja(vendedor, root)

    # si no → comportamiento original
    return ver_detalle_qr(vendedor, root)

def ver_detalle_qr_caja(vendedor_excel, root):
    datos = obtener_datos_caja(vendedor_excel)

    if not datos or datos["qr"] is None or datos["qr"].empty:
        messagebox.showinfo(
            "Sin Datos QR",
            f"No hay transacciones QR para la caja de: {vendedor_excel}"
        )
        return

    # reutilizamos la lógica original
    df_qr = datos["qr"].copy()
    TRANSACCION_QR_BUSCADA_LOCAL = globals().get("TRANSACCION_QR_BUSCADA")

    # ---- desde acá es copia directa de ver_detalle_qr ----

    def texto_comprobante(r):
        nco = r.get("NCO")
        if pd.isna(nco) or str(nco).strip() == "":
            return "SIN COMPROBANTE"
        tip = "" if pd.isna(r.get("TIP")) else str(r.get("TIP")).strip()
        tco = "" if pd.isna(r.get("TCO")) else str(r.get("TCO")).strip()
        return f"{tip} {tco} {str(nco).strip()}".strip()

    df_qr["COMPROBANTE_TXT"] = df_qr.apply(texto_comprobante, axis=1)

    nro_busqueda = TRANSACCION_QR_BUSCADA_LOCAL
    if nro_busqueda:
        df_qr = df_qr[
            df_qr["ID_TRANSACCION"].astype(str) == str(nro_busqueda)
        ]
        if df_qr.empty:
            messagebox.showinfo(
                "Transacción no encontrada",
                "La transacción no se encontró para esta caja."
            )
            return

    top = tk.Toplevel(root)
    top.title(f"Detalle QR (CAJA UNIFICADA): {vendedor_excel}")
    top.geometry("1350x520")
    top.grab_set()

    cols = (
        "Comprobante",
        "Fecha",
        "Importe",
        "Cashout",
        "Descuento",
        "Total QR",
        "Transacción"
    )

    tree = ttk.Treeview(top, columns=cols, show="headings")
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=180, anchor="center")
    tree.pack(fill="both", expand=True, padx=5, pady=5)

    total_interfaz = 0.0
    for _, r in df_qr.iterrows():
        fecha = pd.to_datetime(r["FEC"]).strftime("%d/%m/%Y %H:%M")
        tree.insert(
            "",
            "end",
            values=(
                r["COMPROBANTE_TXT"],
                fecha,
                formatear_moneda_ui(r["IMPORTE"]),
                formatear_moneda_ui(r["CASHOUT"]),
                formatear_moneda_ui(r["DESC_PROMO_NUM"]),
                formatear_moneda_ui(r["QR_FINAL"]),
                r["ID_TRANSACCION"]
            )
        )
        total_interfaz += float(r["QR_FINAL"])

    tk.Label(
        top,
        text=f"TOTAL QR: {formatear_arg(total_interfaz)}",
        font=("Arial", 13, "bold"),
        bg="#2980b9",
        fg="white"
    ).pack(side="bottom", fill="x")

def obtener_datos_caja(vendedor):
    """
    Devuelve TODOS los datos reales de la caja a la que
    pertenece el vendedor.
    """
    w = widgets.get(vendedor)
    if not w:
        return None

    caja_id = w["caja_id"]
    return CAJAS_DATA.get(caja_id)

def ajustar_columnas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def aplicar_header(ws, headers):
    header_fill = PatternFill("solid", fgColor="2c3e50")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    ws.append(headers)

    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = border

    ws.freeze_panes = "A2"
def observaciones_descontar_por_vendedor(vendedor):
    textos = []

    for a in ANOTACIONES_TMP:
        if not son_nombres_similares(vendedor, a.get("vendedor")):
            continue

        tipo = (a.get("tipo") or "").strip().upper()
        if tipo != "DESCONTAR":
            continue

        desc = (a.get("descripcion") or "").strip()
        if desc:
            textos.append(desc)

    return "; ".join(textos)


def observaciones_completas_por_vendedor(vendedor):
    textos = []

    for a in ANOTACIONES_TMP:
        if son_nombres_similares(vendedor, a.get("vendedor")):
            desc = a.get("descripcion", "").strip()
            if desc:
                textos.append(f"* {desc}")

    return "\n".join(textos)
def normalizar_id_transaccion(val):
    if val is None:
        return ""
    s = str(val)
    try:
        if s.lower().startswith("0x"):
            return str(int(s, 16))
    except:
        pass
    return s


def obtener_turno_actual():
    hora = datetime.now().hour

    if 6 <= hora < 15:
        return "Turno_Mañana"
    elif 15 <= hora < 23:
        return "Turno_Tarde"
    else:
        return None  # no se guarda

def guardar_cierre_caja_excel():
    global TURNO_SELECCIONADO, widgets

    if not TURNO_SELECCIONADO:
        messagebox.showwarning("Guardar cierre", "Debe seleccionar un turno.")
        return

    carpeta = "C:/cierres de caja/"
    os.makedirs(carpeta, exist_ok=True)

    fecha = datetime.now().strftime("%Y-%m-%d")
    ruta = os.path.join(
        carpeta,
        f"Cierre_Caja_{TURNO_SELECCIONADO}_{fecha}.xlsx"
    )

    wb = Workbook()

    # =====================================================
    # 🟩 HOJA 1 – RENDICIÓN FINAL POR VENDEDOR
    # =====================================================
    ws_resumen = wb.active
    ws_resumen.title = "RENDICIÓN FINAL POR VENDEDOR"

    headers = ["Planilla", "Vendedor", "Diferencia", "Observaciones"]
    aplicar_header(ws_resumen, headers)

    for v, w in widgets.items():
        diff = parse_moneda_robusto(w["diff"].cget("text"))

        # =========================
        # OBSERVACIONES – SOLO DESCONTAR (PRIMER SOLAPA)
        # =========================
        obs_desc = []

        for a in ANOTACIONES_TMP:
            if not son_nombres_similares(v, a.get("vendedor")):
                continue

            # 👉 DESCONTAR = tiene monto
            monto = a.get("monto")
            if monto in (None, "", 0):
                continue

            desc = (a.get("descripcion") or "").strip()
            if desc:
                obs_desc.append(desc)




        ws_resumen.append([
            PLANILLA_DESDE_SQL or "",
            v,
            diff,
            "; ".join(obs_desc)
        ])


        fila = ws_resumen.max_row
        ws_resumen[f"C{fila}"].font = Font(
            color="c0392b" if diff < 0 else "27ae60",
            bold=True
        )
        ws_resumen[f"C{fila}"].number_format = '"$"#,##0.00;-"$"#,##0.00'

    ajustar_columnas(ws_resumen)

    # =====================================================
    # 🟦 SOLAPA POR VENDEDOR
    # =====================================================
    for v, w in widgets.items():
        ws = wb.create_sheet(v[:31])

        ws["A1"] = "Fecha:"
        ws["B1"] = fecha
        ws["A2"] = "Turno:"
        ws["B2"] = TURNO_SELECCIONADO
        ws["A3"] = "Planilla:"
        ws["B3"] = PLANILLA_DESDE_SQL or ""

        ws["A5"] = "Concepto"
        ws["B5"] = "Importe"
        ws["A5"].font = ws["B5"].font = Font(bold=True)

        fila = 6

        def fila_concepto(nombre, valor, color=None):
            nonlocal fila
            ws[f"A{fila}"] = nombre
            ws[f"B{fila}"] = valor
            ws[f"B{fila}"].number_format = '"$"#,##0.00;-"$"#,##0.00'
            if color:
                ws[f"B{fila}"].font = Font(color=color, bold=True)
            fila += 1

        fila_concepto("Venta Excel", w.get("base_excel", 0))
        fila_concepto("Venta GNC", w.get("gnc_asignado", 0))
        fila_concepto("Total QR", parse_moneda_robusto(w["qr"].get()))
        fila_concepto("Productos Facturados", parse_moneda_robusto(w["prod"].get()))
        fila_concepto("Percepción", parse_moneda_robusto(w["per"].get() or 0))
        # --- NUEVO: Cuentas Corrientes ---
        widget_cta = w.get("cta_cte")
        cta_cte_val = parse_moneda_robusto(widget_cta.get()) if widget_cta else 0.0
        fila_concepto("Cuentas Corrientes", cta_cte_val)
        fila_concepto("Tarjetas Ingresadas", parse_moneda_robusto(w["tarj"].get()))
        fila_concepto("Dinero Rendido", parse_moneda_robusto(w["ef"].get()))
        fila_concepto("Total Anotaciones", total_anotaciones_por_vendedor(v))

        diff = parse_moneda_robusto(w["diff"].cget("text"))
        fila_concepto("Diferencia", diff, "c0392b" if diff < 0 else "27ae60")

        fila += 1
        ws[f"A{fila}"] = "OBSERVACIONES"
        ws[f"A{fila}"].font = Font(bold=True)

        fila += 1
        textos = []
        for a in ANOTACIONES_TMP:
            if son_nombres_similares(v, a.get("vendedor")):
                desc = a.get("descripcion", "").strip()
                if desc:
                    textos.append(f"* {desc}")

        ws.merge_cells(start_row=fila, start_column=1, end_row=fila + max(3, len(textos)), end_column=2)
        ws[f"A{fila}"] = "\n".join(textos) if textos else "—"
        ws[f"A{fila}"].alignment = Alignment(vertical="top", wrap_text=True)

        ajustar_columnas(ws)

    # =====================================================
    # 🟨 DETALLE QR (ORDEN DEFINITIVO)
    # =====================================================
    ws_qr = wb.create_sheet("DETALLE_QR")

    aplicar_header(ws_qr, [
        "Planilla", "Operario", "Fecha", "Comprobante", "Estado",
        "Transacciones", "Importe Neto", "Extracash",
        "Descuentos", "Importe Completo"
    ])

    for df in DATOS_DETALLE_QR.values():
        if df is None or df.empty:
            continue

        for _, r in df.iterrows():
            ws_qr.append([
                PLANILLA_DESDE_SQL or "",
                r.get("Operario"),
                r.get("FEC"),
                f"{r.get('TIP','')}-{r.get('TCO','')}-{r.get('NCO','')}",
                r.get("Estado_QR"),
                normalizar_id_transaccion(r.get("ID_TRANSACCION")),
                r.get("IMPORTE"),
                r.get("CASHOUT"),
                r.get("DESC_PROMO"),
                r.get("QR_FINAL")
            ])

    ajustar_columnas(ws_qr)

    # =====================================================
    # 🟧 DETALLE PRODUCTOS FACTURADOS
    # =====================================================
    ws_prod = wb.create_sheet("DETALLE_PRODUCTOS_FACTURADOS")

    if not DATOS_DETALLE_FACTURACION:
        ws_prod.append(["Sin productos facturados"])
    else:
        headers_prod = next(
            (list(df.columns) for df in DATOS_DETALLE_FACTURACION.values() if df is not None and not df.empty),
            None
        )

        if headers_prod:
            aplicar_header(ws_prod, headers_prod)
            for df in DATOS_DETALLE_FACTURACION.values():
                if df is None or df.empty:
                    continue
                for _, row in df.iterrows():
                    ws_prod.append(list(row.values))

    ajustar_columnas(ws_prod)
    # =====================================================
    # 🟪 DETALLE CTAS CTES (REMITOS)
    # =====================================================
    ws_cta = wb.create_sheet("DETALLE_CTAS_CTES")

    if not DATOS_DETALLE_REMITOS:
        ws_cta.append(["Sin cuentas corrientes registradas"])
    else:
        headers_cta = ["Fecha", "Tipo", "Comprobante", "Cliente", "Total", "Vendedor"]
        aplicar_header(ws_cta, headers_cta)

        for df in DATOS_DETALLE_REMITOS.values():
            if df is None or df.empty:
                continue
            for _, r in df.iterrows():
                # Formatear la fecha para Excel
                f_str = pd.to_datetime(r['Fecha']).strftime("%d/%m/%Y %H:%M") if pd.notna(r['Fecha']) else ""
                # Validar el cliente
                cli = str(r['Cliente']).strip() if pd.notna(r['Cliente']) and str(r['Cliente']).strip() != "" else "CONSUMIDOR FINAL"

                ws_cta.append([
                    f_str,
                    r.get('Tipo', 'RE'),
                    r.get('Comprobante', ''),
                    cli,
                    float(r.get('Total_Remito', 0.0)),
                    r.get('Nombre_Vendedor', '')
                ])
    ajustar_columnas(ws_cta)

    # -----------------------------------------------------
    # FIN DEL ARMADO DEL EXCEL
    # -----------------------------------------------------

    wb.save(ruta)

    messagebox.showinfo(
        "Cierre guardado",
        f"Cierre guardado correctamente:\n{ruta}"
    )


def on_cerrar_programa(root):
    respuesta = messagebox.askyesnocancel(
        "Cerrar programa",
        "El cierre de caja no fue guardado.\n\n¿Desea guardarlo ahora?"
    )

    if respuesta is True:
        # ✅ Sí → abrir ventana de guardado
        abrir_ventana_guardado(root)

    # ❌ NO y ❌ Cancelar → no hacer absolutamente nada
    return




def abrir_ventana_guardado(root):
    global TURNO_SELECCIONADO

    FECHA_ACTUAL = datetime.now().strftime("%d/%m/%Y")

    top = tk.Toplevel(root)
    top.title("Guardar cierre de caja")
    top.geometry("360x260")
    top.resizable(False, False)

    # -------------------------
    # Modal real + bloqueo X
    # -------------------------
    top.transient(root)
    top.grab_set()

    def bloquear_cierre():
        messagebox.showwarning(
            "Cierre obligatorio",
            "Debe guardar el cierre de caja antes de salir."
        )

    top.protocol("WM_DELETE_WINDOW", bloquear_cierre)

    # =========================
    # CONTENIDO
    # =========================
    cont = tk.Frame(top)
    cont.pack(fill="both", expand=True, padx=15, pady=15)

    # -------- Fecha --------
    tk.Label(cont, text="Fecha").pack(anchor="w")

    entry_fecha = tk.Entry(cont)
    entry_fecha.insert(0, FECHA_ACTUAL)
    entry_fecha.config(
        state="disabled",
        disabledforeground="#7f8c8d",
        disabledbackground="#ecf0f1"
    )
    entry_fecha.pack(fill="x", pady=(0, 12))

    # -------- Planilla --------
    tk.Label(cont, text="Número de planilla").pack(anchor="w")

    entry_planilla = tk.Entry(cont)
    if PLANILLA_DESDE_SQL and PLANILLA_HASTA_SQL:
        texto_planilla = f"{PLANILLA_DESDE_SQL} → {PLANILLA_HASTA_SQL}"
    else:
        texto_planilla = "SIN CONSULTA SQL"

    entry_planilla.insert(0, texto_planilla)

    entry_planilla.config(
        state="disabled",
        disabledforeground="#7f8c8d",
        disabledbackground="#ecf0f1"
    )
    entry_planilla.pack(fill="x", pady=(0, 12))

    # -------- Turno (Combo) --------
    tk.Label(cont, text="Turno").pack(anchor="w")

    turno_var = tk.StringVar()

    combo_turno = ttk.Combobox(
        cont,
        textvariable=turno_var,
        state="readonly",
        values=["Turno Mañana", "Turno Tarde"]
    )
    combo_turno.pack(fill="x", pady=(0, 20))

    # -------- Botón Guardar --------
    def confirmar_guardado():
        global TURNO_SELECCIONADO

        if not turno_var.get():
            messagebox.showwarning(
                "Turno requerido",
                "Debe seleccionar un turno antes de guardar."
            )
            return

        TURNO_SELECCIONADO = turno_var.get()

        guardar_cierre_caja_excel()

        # 🔴 cerrar TODA la app
        root.destroy()

    tk.Button(
        cont,
        text="Guardar cierre",
        bg="#27ae60",
        fg="white",
        font=("Arial", 10, "bold"),
        command=confirmar_guardado
    ).pack(fill="x")
    btns = tk.Frame(cont)
    btns.pack(fill="x")

    tk.Button(
        btns,
        text="Cancelar",
        bg="#bdc3c7",
        fg="black",
        font=("Arial", 10, "bold"),
        command=lambda: top.destroy()
    ).pack(side="left", expand=True, fill="x", padx=(0, 5))


def eliminar_transaccion_seleccionada():
    global ANOTACIONES_TMP, TREE_ANOTACIONES

    if TREE_ANOTACIONES is None or not TREE_ANOTACIONES.winfo_exists():
        return

    seleccion = TREE_ANOTACIONES.selection()
    if not seleccion:
        messagebox.showwarning(
            "Eliminar transacción",
            "Seleccione una transacción para eliminar."
        )
        return

    item_id = seleccion[0]
    valores = TREE_ANOTACIONES.item(item_id, "values")

    nro_transaccion = str(valores[0]).strip()

    confirmar = messagebox.askyesno(
        "Eliminar transacción",
        f"¿Eliminar la transacción {nro_transaccion}?\n\n"
        "⚠️ Esta acción NO se puede deshacer."
    )

    if not confirmar:
        return

    # -------------------------
    # Eliminar de ANOTACIONES_TMP
    # -------------------------
    ANOTACIONES_TMP = [
        a for a in ANOTACIONES_TMP
        if str(a.get("transaccion")) != nro_transaccion
    ]

    # -------------------------
    # Limpiar QR reasignado (si existía)
    # -------------------------
    if nro_transaccion in QR_REASIGNADOS:
        QR_REASIGNADOS.remove(nro_transaccion)

    # -------------------------
    # Refrescar TODO lo relacionado a anotaciones
    # -------------------------
    refrescar_anotaciones_ui()
    actualizar_labels_anotaciones()
    refrescar_calculo_principal()

    messagebox.showinfo(
        "Transacción eliminada",
        f"La transacción {nro_transaccion} fue eliminada correctamente."
    )



def normalizar_desc_promo(texto):
    if not texto:
        return 0.0

    s = str(texto)

    # Captura números con miles y decimales en ambos formatos
    nums = re.findall(r"\d[\d.,]*", s)
    if not nums:
        return 0.0

    total = 0.0

    for n in nums:
        original = n

        # 🇦🇷 Formato argentino: 4.800,00
        if "." in n and "," in n and n.rfind(".") < n.rfind(","):
            n = n.replace(".", "").replace(",", ".")
        # 🇺🇸 Formato USA: 4,800.00
        elif "." in n and "," in n and n.rfind(",") < n.rfind("."):
            n = n.replace(",", "")
        # Miles sin decimales: 4.800 o 4,800
        elif n.count(".") == 1 and len(n.split(".")[1]) == 3:
            n = n.replace(".", "")
        elif n.count(",") == 1 and len(n.split(",")[1]) == 3:
            n = n.replace(",", "")
        # Decimal con coma
        elif "," in n:
            n = n.replace(",", ".")

        try:
            total += float(n)
        except ValueError:
            pass

    return round(total, 2)



def limpiar_texto_monetario(valor):
    """
    Extrae el número (con signo) de un texto.
    Ejemplos:
    '$1.200 promo' -> '1.200'
    'promo -500'   -> '-500'
    'DESC $300'    -> '300'
    """
    if valor is None:
        return ""
    s = str(valor)
    # deja solo dígitos, coma, punto y signo -
    limpio = re.findall(r'-?\d+[.,]?\d*', s)
    return limpio[0] if limpio else ""
#--------------------prueba----------------------------------
def simular_qr_transaccion_prueba():
    import pandas as pd
    from random import randint

    global DATOS_DETALLE_QR

    nro_tx = "999999"
    nro_comprobante = f"B FT {randint(10000, 99999)}"

    fila_fake = {
        "ID_TRANSACCION": nro_tx,
        "QR_FINAL": 20000.00,
        "IMPORTE": 20000.00,
        "VENDEDOR": "MARCOS",
        "Operario": "MARCOS",
        "NCO": nro_comprobante,     # ✅ comprobante REAL
        "FECHA": "2026-01-13",
        "HORA": "13:00"
    }

    df_fake = pd.DataFrame([fila_fake])

    if "MARCOS" not in DATOS_DETALLE_QR or DATOS_DETALLE_QR["MARCOS"] is None:
        DATOS_DETALLE_QR["MARCOS"] = df_fake
    else:
        DATOS_DETALLE_QR["MARCOS"] = pd.concat(
            [DATOS_DETALLE_QR["MARCOS"], df_fake],
            ignore_index=True
        )

    messagebox.showinfo(
        "QR SIMULADO",
        f"Se simuló la transacción {nro_tx}\nComprobante: {nro_comprobante}"
    )

#--------------------------------------------------------------
def actualizar_completo():
    if EJECUTAR_SQL_HANDLER is None:
        return

    if PLANILLA_DESDE_SQL is None or PLANILLA_HASTA_SQL is None:
        return  # no hace nada, como pediste

    EJECUTAR_SQL_HANDLER(
        PLANILLA_DESDE_SQL,
        PLANILLA_HASTA_SQL
    )

    actualizar_anotaciones_y_qr()
    refrescar_calculo_principal()


def inicializar_vendedor_desde_sql(vendedor):
    """
    Inyecta QR, productos, percepción y efectivo SQL
    en un vendedor creado dinámicamente (ej: GNC).
    """
    w = widgets.get(vendedor)
    if not w:
        return

    # =========================
    # QR
    # =========================
    total_qr = 0.0
    for vend_sql, df in DATOS_DETALLE_QR.items():
        if son_nombres_similares(vendedor, vend_sql):
            total_qr = df["QR_FINAL"].sum()
            break

    w["qr_db"] = total_qr
    w["qr"].config(state="normal")
    actual = parse_moneda_robusto(w["qr"].get())
    w["qr"].delete(0, tk.END)
    w["qr"].insert(0, f"{total_qr:.2f}")
    w["qr"].config(state="disabled")

    # =========================
    # PRODUCTOS FACTURADOS
    # =========================
    for vend_sql, df in DATOS_DETALLE_FACTURACION.items():
        if son_nombres_similares(vendedor, vend_sql):
            total_prod = df["Total_Neto"].sum()
            w["prod"].delete(0, tk.END)
            w["prod"].insert(0, f"{total_prod:.2f}")
            w["prod"].config(state="disabled")
            break

    # =========================
    # PERCEPCIÓN
    # =========================
    if DF_RENDICIONES_CACHE is not None:
        for _, r in DF_RENDICIONES_CACHE.iterrows():
            if son_nombres_similares(vendedor, r.get("Vendedor", "")):
                monto = parse_moneda_robusto(r.get("Percepcion", 0))
                w["per"].delete(0, tk.END)
                w["per"].insert(0, f"{monto:.2f}")
                w["per"].config(state="disabled")
                break

    # =========================
    # EFECTIVO (+ y lápiz)
    # =========================
    # habilita lógica normal de efectivo
    if vendedor not in DATOS_RENDICIONES:
        DATOS_RENDICIONES[vendedor] = {"movimientos": []}

    refrescar_efectivos_ui(vendedor, w)

def inyectar_gnc_a_responsable(responsable_ui):
    """
    Inyecta movimientos GNC SQL a un responsable
    aunque haya sido agregado DESPUÉS de consultar SQL.
    """
    global DF_GNC_SQL_CACHE, DATOS_RENDICIONES

    if DF_GNC_SQL_CACHE is None or DF_GNC_SQL_CACHE.empty:
        return

    for _, row in DF_GNC_SQL_CACHE.iterrows():
        vendedor_sql = str(row.get("Vendedor", "")).strip().upper()

        if son_nombres_similares(responsable_ui, vendedor_sql):

            DATOS_RENDICIONES \
                .setdefault(responsable_ui, {}) \
                .setdefault("movimientos", []) \
                .append({
                    "origen": "sql",
                    "tipo": "GNC",
                    "planilla": row.get("Planilla"),
                    "nro": row.get("Nro_Mov"),
                    "ref": f"GNC {row.get('Nro_Mov')} - {row.get('Fecha'):%d/%m %H:%M}",
                    "monto": float(parse_moneda_robusto(row.get("Efectivo", 0)))
                })
def total_efectivo_vendedor(vendedor):
    movimientos = DATOS_RENDICIONES.get(vendedor, {}).get("movimientos", [])

    return sum(
        parse_moneda_robusto(m["monto"])
        for m in movimientos
        if m.get("origen") in ("sql", "manual")
    )


def ventana_asignar_qr_sin_comprobante(root, nro_transaccion, fila_qr):
    global ANOTACIONES_TMP

    top = tk.Toplevel(root)
    top.title("Asignar QR sin comprobante")
    top.geometry("460x300")
    top.resizable(False, False)
    top.grab_set()

    tk.Label(
        top,
        text=f"QR SIN COMPROBANTE\nTransacción {nro_transaccion}",
        font=("Arial", 11, "bold"),
        fg="#e67e22",
        justify="center"
    ).pack(pady=10)

    monto_qr = float(fila_qr.get("QR_FINAL", 0.0))

    # -------------------------
    # Monto
    # -------------------------
    tk.Label(top, text="Monto").pack(anchor="w", padx=12)
    entry_monto = tk.Entry(top, justify="right")
    entry_monto.insert(0, formatear_arg(monto_qr))
    entry_monto.config(state="disabled")
    entry_monto.pack(padx=12, pady=(0, 10))

    # -------------------------
    # Vendedor
    # -------------------------
    tk.Label(top, text="Asignar a vendedor").pack(anchor="w", padx=12)

    cargar_vendedores_db()

    vendedores = (
        DF_VENDEDORES_CACHE["NomVen"]
        .dropna()
        .str.strip()
        .str.upper()
        .unique()
        .tolist()
    )

    frame_auto, var_vendedor = crear_autocomplete(top, vendedores)
    frame_auto.pack(padx=12, pady=10, anchor="w")

    # -------------------------
    # Confirmar
    # -------------------------
    def confirmar_asignacion():
        vendedor = var_vendedor.get().strip().upper()

        if not vendedor or vendedor not in vendedores:
            messagebox.showwarning(
                "Vendedor inválido",
                "Seleccione un vendedor válido."
            )
            return

        # 🔴 AGREGAR COMO ANOTACIÓN NORMAL
        anotacion = {
            "vendedor": normalizar_texto(vendedor),
            "monto": monto_qr,
            "descripcion": f"QR SIN COMPROBANTE ({nro_transaccion})",
            "tipo": "QR",
            "estado": ESTADO_QR_SIN_COMPROBANTE,
            "transaccion": nro_transaccion,
            "fila_qr": None
        }

        ANOTACIONES_TMP.append(anotacion)
        # 🔥 SACAR EL QR DEL CIRCUITO QR
        for operario, df in DATOS_DETALLE_QR.items():
            if df is None or df.empty:
                continue

            DATOS_DETALLE_QR[operario] = df[
                df["ID_TRANSACCION"].astype(str) != str(nro_transaccion)
            ]


        # 🔴 REFRESCAR TODO
        refrescar_anotaciones_ui()
        refrescar_calculo_principal()
        actualizar_labels_anotaciones()

        messagebox.showinfo(
            "QR asignado",
            f"El QR {nro_transaccion} fue asignado a {vendedor}."
        )

        top.destroy()

    tk.Button(
        top,
        text="Confirmar asignación",
        bg="#27ae60",
        fg="white",
        font=("Arial", 10, "bold"),
        width=22,
        command=confirmar_asignacion
    ).pack(pady=10)

    tk.Button(
        top,
        text="Cancelar",
        width=18,
        command=top.destroy
    ).pack()

def ventana_asignar_qr_sin_comprobante(root, nro_transaccion, fila_qr):
    """
    Permite asignar un QR SIN COMPROBANTE a un vendedor
    y lo agrega directamente a Anotaciones.
    """
    top = tk.Toplevel(root)
    top.title("QR sin comprobante")
    top.geometry("420x220")
    top.resizable(False, False)
    top.grab_set()

    tk.Label(
        top,
        text=f"Transacción QR {nro_transaccion}",
        font=("Arial", 11, "bold"),
        fg="#c0392b"
    ).pack(pady=(10, 5))

    tk.Label(
        top,
        text="La transacción no tiene comprobante.\n¿A qué vendedor desea asignarla?",
        justify="center"
    ).pack(pady=5)

    # -------------------------
    # Vendedor (autocomplete DB)
    # -------------------------
    cargar_vendedores_db()

    lista_vendedores = sorted(
        DF_VENDEDORES_CACHE["NomVen"]
        .dropna()
        .str.strip()
        .str.upper()
        .tolist()
    )

    frame_auto, var_vendedor = crear_autocomplete(top, lista_vendedores)
    frame_auto.pack(pady=10)

    # -------------------------
    # Botones
    # -------------------------
    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)

    def confirmar():
        vendedor = var_vendedor.get().strip().upper()

        if vendedor not in lista_vendedores:
            messagebox.showwarning(
                "Vendedor inválido",
                "Seleccione un vendedor válido."
            )
            return

        monto = float(fila_qr.get("QR_FINAL", 0.0))

        agregar_anotacion(
            nro_transaccion,
            vendedor,
            monto,
            "QR sin comprobante (asignado manualmente)"
        )

        # 🔴 NO marcar como impacto QR
        ANOTACIONES_TMP[-1]["estado"] = ESTADO_QR_ASIGNADO_MANUAL
        ANOTACIONES_TMP[-1]["origen_qr"] = ESTADO_QR_SIN_COMPROBANTE
        ANOTACIONES_TMP[-1]["qr_fila"] = None




        refrescar_calculo_principal()
        refrescar_anotaciones_ui()
        actualizar_labels_anotaciones()

        top.destroy()

    tk.Button(
        frame_btn,
        text="Asignar",
        bg="#27ae60",
        fg="white",
        width=14,
        command=confirmar
    ).pack(side="left", padx=5)

    tk.Button(
        frame_btn,
        text="Cancelar",
        width=14,
        command=top.destroy
    ).pack(side="left", padx=5)


def extraer_numeros(texto):
    if not texto:
        return []
    return re.findall(r'\d+', str(texto))

def aplicar_qr_a_vendedor_desde_anotacion(anot):
    """
    Impacta el QR_FINAL de una anotación al vendedor,
    evitando duplicados.
    """
    qr_fila = anot.get("qr_fila")

    if qr_fila is None or not isinstance(qr_fila, pd.Series):
        return


    nro = str(anot.get("transaccion"))

    if nro in QR_REASIGNADOS:
        return

    fila = anot["qr_fila"]
    monto = float(fila.get("QR_FINAL", 0.0))
    vendedor = anot["vendedor"]

    if monto == 0:
        return

    DATOS_RENDICIONES.setdefault(vendedor, {}) \
        .setdefault("movimientos", []) \
        .append({
            "origen": "qr_anotacion",
            "tipo": "QR reasignado",
            "ref": f"QR {nro}",
            "monto": monto
        })

    QR_REASIGNADOS.add(nro)

def on_click_anotaciones(event):
    global TREE_ANOTACIONES

    if TREE_ANOTACIONES is None:
        return

    if not TREE_ANOTACIONES.winfo_exists():
        return

    # Identificar fila y columna
    item_id = TREE_ANOTACIONES.identify_row(event.y)
    col = TREE_ANOTACIONES.identify_column(event.x)

    if not item_id or col != "#5":  # columna 5 = VER QR
        return

    valores = TREE_ANOTACIONES.item(item_id, "values")
    nro_transaccion = str(valores[0]).strip()

    # Buscar la anotación correspondiente
    for a in ANOTACIONES_TMP:
        if str(a["transaccion"]) == nro_transaccion:
            # Solo si impactó en QR
            if a.get("estado") == ESTADO_QR_IMPACTADO:

                vendedor_qr = a.get("qr_fila").get("Operario") if "Operario" in a.get("qr_fila") else None
                if vendedor_qr:
                    # Usamos la misma ventana principal (vent) para mostrar detalle
                    ver_detalle_qr(vendedor_qr, TREE_ANOTACIONES.master)
            break

def actualizar_anotaciones_y_qr():
    """
    Reconsulta QR, cruza contra anotaciones
    y refresca TODA la UI.
    """

    if CALCULAR_HANDLER:
        CALCULAR_HANDLER()

    cambio = reconciliar_anotaciones_con_qr()

    if cambio:
        actualizar_labels_anotaciones()
        refrescar_anotaciones_ui()
    # 2️⃣ Cruzar anotaciones con QR
    reconciliar_anotaciones_con_qr()

    # 3️⃣ Refrescar columna de anotaciones
    actualizar_labels_anotaciones()

    # 4️⃣ Refrescar ventana de anotaciones (si está abierta)
    refrescar_anotaciones_ui()
def reconciliar_anotaciones_con_qr():
    """
    Cruza anotaciones PENDIENTES contra QR reales.
    SOLO impacta:
    - si el QR existe
    - si tiene comprobante
    - si aún no fue impactado
    """

    if not DATOS_DETALLE_QR:
        return False

    hubo_cambios = False

    for anot in ANOTACIONES_TMP:

        if anot.get("estado") != ESTADO_QR_PENDIENTE:
            continue


        nro = str(anot.get("transaccion", "")).strip()
        if not nro:
            continue

        estado, vendedor_qr, fila = buscar_transaccion_qr(nro)

        if estado != "OK" or fila is None:
            continue



        # 🔴 validar operario
        if not vendedor_qr:
            continue

        # ✅ IMPACTO REAL
        anot["estado"] = ESTADO_QR_IMPACTADO
        anot["qr_fila"] = fila



        aplicar_qr_a_vendedor_desde_anotacion(anot)


        hubo_cambios = True

    return hubo_cambios

def refrescar_anotaciones_ui():
    global TREE_ANOTACIONES, ANOTACIONES_TMP

    if TREE_ANOTACIONES is None or not TREE_ANOTACIONES.winfo_exists():
        return

    TREE_ANOTACIONES.delete(*TREE_ANOTACIONES.get_children())

    for a in ANOTACIONES_TMP:

        estado = a.get("estado", "")
        origen = a.get("origen_qr", "")

        if estado == ESTADO_QR_IMPACTADO:
            estado_qr = "IMPACTÓ QR"
            tag = "impacto_qr"

        elif estado == ESTADO_QR_MANUAL:
            estado_qr = "ASIGNADO MANUAL"
            tag = "qr_asignado"
        elif estado == ESTADO_QR_ASIGNADO_MANUAL:
            estado_qr = "ASIGNADO MANUAL"
            tag = "qr_asignado"


        elif estado == ESTADO_QR_PENDIENTE:
            if origen == ESTADO_QR_SIN_COMPROBANTE:
                estado_qr = "PENDIENTE (SIN COMPROBANTE)"
            else:
                estado_qr = "PENDIENTE"
            tag = "no_encontrada"

        else:
            estado_qr = estado or "NO ENCONTRADA"
            tag = "no_encontrada"


        TREE_ANOTACIONES.insert(
            "",
            "end",
            values=(
                a.get("transaccion", ""),
                a.get("vendedor", ""),
                formatear_moneda_ui(a.get("monto", 0)),
                a.get("descripcion", ""),
                estado_qr
            ),
            tags=(tag,)
        )


def ventana_tipo_anotacion(root):
    top = tk.Toplevel(root)
    top.title("Tipo de anotación")
    top.geometry("360x180")
    top.resizable(False, False)
    top.grab_set()

    tk.Label(
        top,
        text="¿Qué tipo de anotación desea crear?",
        font=("Arial", 11, "bold")
    ).pack(pady=15)

    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)

    # ℹ️ INFORMAR (no impacta en cálculos)
    tk.Button(
        frame_btn,
        text="ℹ️ Informar",
        bg="#3498db",
        fg="white",
        font=("Arial", 9, "bold"),
        width=18,
        command=lambda: (
            top.destroy(),
            ventana_informar(root)
        )
    ).pack(pady=5)

    # ➖ DESCONTAR (impacta en salida)
    tk.Button(
        frame_btn,
        text="➖ Descontar",
        bg="#c0392b",
        fg="white",
        font=("Arial", 9, "bold"),
        width=18,
        command=lambda: (
            top.destroy(),
            ventana_descontar(root)
        )
    ).pack(pady=5)
def ventana_descontar(root):
    top = tk.Toplevel(root)
    top.title("Descontar anotación")
    top.geometry("520x380")
    top.resizable(False, False)
    top.grab_set()

    tk.Label(
        top,
        text="Anotación para descontar",
        font=("Arial", 11, "bold"),
        fg="#c0392b"
    ).pack(pady=8)

    # -------------------------
    # Nro transacción
    # -------------------------
    tk.Label(top, text="Número de transacción").pack(anchor="w", padx=12)
    entry_tx = tk.Entry(top, width=30)
    entry_tx.pack(padx=12, pady=(0, 8))

    # -------------------------
    # Monto
    # -------------------------
    tk.Label(top, text="Monto a descontar").pack(anchor="w", padx=12)
    entry_monto = tk.Entry(top, width=20, justify="right")
    entry_monto.pack(padx=12, pady=(0, 8))

    # -------------------------
    # Descripción
    # -------------------------
    tk.Label(top, text="Descripción").pack(anchor="w", padx=12)


    txt_desc = tk.Text(
        top,
        width=60,
        height=6,
        wrap="word"
    )
    txt_desc.pack(padx=12, pady=(0, 8))
    # -------------------------
    # Vendedor
    # -------------------------
    tk.Label(top, text="Vendedor / Operario").pack(anchor="w", padx=12)

    cargar_vendedores_db()
    lista_vendedores = sorted(
        DF_VENDEDORES_CACHE["NomVen"]
        .dropna()
        .str.strip()
        .str.upper()
        .tolist()
    )

    frame_auto, var_vendedor = crear_autocomplete(top, lista_vendedores)
    frame_auto.pack(padx=12, pady=(0, 12), anchor="w")

    # -------------------------
    # Botones
    # -------------------------
    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)

    def confirmar():
        nro = entry_tx.get().strip()
        desc = txt_desc.get("1.0", "end").strip()
        vendedor = var_vendedor.get().strip().upper()
        monto_txt = entry_monto.get().strip()

        if not nro or not desc or not vendedor or not monto_txt:
            messagebox.showwarning(
                "Datos incompletos",
                "Debe completar todos los campos."
            )
            return

        try:
            monto = float(monto_txt.replace(",", "."))
        except ValueError:
            messagebox.showwarning(
                "Monto inválido",
                "Ingrese un monto válido."
            )
            return

        agregar_anotacion(
            nro,
            vendedor,
            monto,
            desc
        )

        refrescar_calculo_principal()

        messagebox.showinfo(
            "Anotación descontada",
            "La anotación fue aplicada a la salida."
        )

        top.destroy()

    tk.Button(
        frame_btn,
        text="Confirmar",
        bg="#c0392b",
        fg="white",
        width=15,
        command=confirmar
    ).pack(side="left", padx=5)

    tk.Button(
        frame_btn,
        text="Cancelar",
        width=15,
        command=top.destroy
    ).pack(side="left", padx=5)

def ventana_informar(root):
    top = tk.Toplevel(root)
    top.title("Informar anotación")
    top.geometry("520x300")
    top.resizable(False, False)
    top.grab_set()

    tk.Label(
        top,
        text="Anotación informativa",
        font=("Arial", 11, "bold"),
        fg="#2980b9"
    ).pack(pady=8)

    # -------------------------
    # Descripción
    # -------------------------
    tk.Label(top, text="Descripción").pack(anchor="w", padx=12)

    txt_desc = tk.Text(
        top,
        width=60,
        height=6,
        wrap="word"
    )
    txt_desc.pack(padx=12, pady=(0, 8))


    # -------------------------
    # Vendedor
    # -------------------------
    tk.Label(top, text="Vendedor / Operario").pack(anchor="w", padx=12)

    cargar_vendedores_db()
    lista_vendedores = sorted(
        DF_VENDEDORES_CACHE["NomVen"]
        .dropna()
        .str.strip()
        .str.upper()
        .tolist()
    )

    frame_auto, var_vendedor = crear_autocomplete(top, lista_vendedores)
    frame_auto.pack(padx=12, pady=(0, 12), anchor="w")

    # -------------------------
    # Botones
    # -------------------------
    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)

    def confirmar():
        desc = txt_desc.get("1.0", "end").strip()

        vendedor = var_vendedor.get().strip().upper()

        if not desc or not vendedor:
            messagebox.showwarning(
                "Datos incompletos",
                "Debe completar todos los campos."
            )
            return

        agregar_anotacion(
            "INFORMAR",
            vendedor,
            0.0,
            desc
        )

        messagebox.showinfo(
            "Anotación informada",
            "La anotación fue registrada solo a modo informativo."
        )

        top.destroy()

    tk.Button(
        frame_btn,
        text="Confirmar",
        bg="#2980b9",
        fg="white",
        width=15,
        command=confirmar
    ).pack(side="left", padx=5)

    tk.Button(
        frame_btn,
        text="Cancelar",
        width=15,
        command=top.destroy
    ).pack(side="left", padx=5)

def refrescar_calculo_principal():
    """
    Ejecuta el cálculo real de la planilla si existe.
    """
    if CALCULAR_HANDLER:
        CALCULAR_HANDLER()

    cambio = reconciliar_anotaciones_con_qr()

    if cambio:
        actualizar_labels_anotaciones()
        refrescar_anotaciones_ui()
        actualizar_anotaciones_y_qr()
        # 🔹 Actualiza diferencias visuales
    if ACTUALIZAR_DIFERENCIAS_FN:
        ACTUALIZAR_DIFERENCIAS_FN()

    # 🔹 Actualiza total caja conjunta
    if TOTAL_CAJA_CONJUNTA_FN:
        TOTAL_CAJA_CONJUNTA_FN()
# Llamada a la función global
if TOTAL_CAJA_CONJUNTA_FN:
    TOTAL_CAJA_CONJUNTA_FN()
def observaciones_descuentos_por_vendedor(vendedor):
    textos = []

    for a in ANOTACIONES_TMP:
        if not son_nombres_similares(vendedor, a.get("vendedor")):
            continue

        # 🔴 DESCONTAR = tiene monto
        monto = a.get("monto")
        if monto in (None, "", 0):
            continue

        desc = a.get("descripcion", "").strip()
        if desc:
            textos.append(desc)

    return "; ".join(textos)


def total_anotaciones_por_vendedor(vendedor):
    total = 0.0
    for a in ANOTACIONES_TMP:
        if a.get("estado") == "IMPACTO_QR":
            continue
        if a.get("estado") == ESTADO_QR_IMPACTADO:
            continue


        if son_nombres_similares(vendedor, a["vendedor"]):
            total += float(a["monto"])
    return total

def actualizar_lista():
    actualizar_labels_anotaciones()
    refrescar_efectivos_ui()


def agregar_anotacion(transaccion, vendedor, monto, descripcion):
    """
    Agrega una anotación en memoria y refresca la UI correctamente.
    """
    # 🔹 normalizar vendedor EXACTAMENTE igual que la planilla
    vend_norm = normalizar_texto(vendedor)

    ANOTACIONES_TMP.append({
        "transaccion": str(transaccion).strip(),
        "vendedor": vend_norm,
        "monto": float(monto),
        "descripcion": descripcion.strip(),
        "tipo": "QR",

        # ✅ SIEMPRE PENDIENTE
        "estado": ESTADO_QR_PENDIENTE,

        # 🔍 origen informativo (no bloquea)
        "origen_qr": ESTADO_QR_SIN_COMPROBANTE,

        "qr_fila": None
    })



    refrescar_anotaciones_ui()
    # 🔹 actualizar SOLO la columna Anotaciones
    actualizar_labels_anotaciones()

    # 🔹 recalcular TODA la planilla (entrada / salida / diferencia)
    if CALCULAR_HANDLER:
        CALCULAR_HANDLER()

    cambio = reconciliar_anotaciones_con_qr()

    if cambio:
        actualizar_labels_anotaciones()
        refrescar_anotaciones_ui()
def actualizar_labels_anotaciones():
    for vendedor, lbl in LABELS_ANOTACIONES.items():
        total = total_anotaciones_por_vendedor(vendedor)
        lbl.config(text=formatear_arg(total))


def abrir_anotaciones(root):
    global TREE_ANOTACIONES
    global VENTANA_ANOTACIONES

    # 🔁 Si ya existe, traer al frente
    if VENTANA_ANOTACIONES and VENTANA_ANOTACIONES.winfo_exists():
        VENTANA_ANOTACIONES.deiconify()
        VENTANA_ANOTACIONES.lift()
        VENTANA_ANOTACIONES.focus_force()
        return

    reconciliar_anotaciones_con_qr()

    top = tk.Toplevel(root)
    VENTANA_ANOTACIONES = top

    top.title("Anotaciones")
    top.geometry("700x460")

    top.transient(root)
    top.lift()
    top.focus_force()

    # 🔹 nada de topmost permanente

    def auto_actualizar():
        if TREE_ANOTACIONES and TREE_ANOTACIONES.winfo_exists():
            actualizar_anotaciones_y_qr()
            top.after(120_000, auto_actualizar)

    auto_actualizar()

    def on_close():
        global VENTANA_ANOTACIONES
        VENTANA_ANOTACIONES = None
        top.destroy()

    top.protocol("WM_DELETE_WINDOW", on_close)

    # -------------------------
    # Tabla
    # -------------------------
    cols = ("Transacción", "Vendedor", "Monto", "Descripción", "estado_QR")
    TREE_ANOTACIONES = ttk.Treeview(top, columns=cols, show="headings")
    TREE_ANOTACIONES.bind("<Button-1>", on_click_anotaciones)

    for c in cols:
        TREE_ANOTACIONES.heading(c, text=c)
        TREE_ANOTACIONES.column(
            c,
            anchor="w" if c == "Descripción" else "center",
            width=260 if c == "Descripción" else 120
        )

    TREE_ANOTACIONES.pack(fill="both", expand=True, padx=5, pady=5)
    refrescar_anotaciones_ui()


    # -------------------------
    # Botonera inferior
    # -------------------------
    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=8)

    # ➕ Crear anotación manual
    tk.Button(
        frame_btn,
        text="➕ Crear anotación",
        bg="#27ae60",
        fg="white",
        font=("Arial", 9, "bold"),
        width=18,
        command=lambda: ventana_tipo_anotacion(root)
    ).pack(side="left", padx=5)

    # 🔍 Buscar transacción (flujo actual)
    tk.Button(
        frame_btn,
        text="🔍 Buscar transacción",
        bg="#2980b9",
        fg="white",
        font=("Arial", 9, "bold"),
        width=18,
        command=lambda: accion_buscar_transaccion(root)
    ).pack(side="left", padx=5)
    tk.Button(
        frame_btn,
        text="🗑️Eliminar transacción",
        bg="#c0392b",
        fg="white",
        font=("Arial", 9, "bold"),
        width=22,
        command=eliminar_transaccion_seleccionada
    ).pack(side="left", padx=5)

    btn_actualizar = ttk.Button(
        top,
        text="🔄 Actualizar",
        command=actualizar_anotaciones_y_qr
    )
    btn_actualizar.pack(pady=5)

        # ❌ Cerrar
    tk.Button(
        frame_btn,
        text="Cerrar",
        width=12,
        command=top.destroy
    ).pack(side="left", padx=5)
    top.after(120_000,auto_actualizar)



def ventana_anotacion_visual(nro_transaccion, root):
    """
    Ventana VISUAL para transacción no encontrada.
    NO guarda en DB.
    NO afecta conciliación.
    Guarda solo en memoria (ANOTACIONES_TMP).
    """

    top = tk.Toplevel(root)
    top.title("Transacción no encontrada")
    top.geometry("520x360")
    top.resizable(False, False)
    top.grab_set()

    # -------------------------
    # Título
    # -------------------------
    tk.Label(
        top,
        text=f"Transacción: {nro_transaccion}",
        font=("Arial", 11, "bold"),
        fg="#c0392b"
    ).pack(pady=(10, 5))

    # -------------------------
    # Descripción
    # -------------------------
    tk.Label(top, text="Descripción").pack(anchor="w", padx=12)
    entry_desc = tk.Entry(top, width=55)
    entry_desc.pack(padx=12, pady=(0, 8))

    # -------------------------
    # Monto
    # -------------------------
    tk.Label(top, text="Monto de la operación").pack(anchor="w", padx=12)
    entry_monto = tk.Entry(top, width=20, justify="right")
    entry_monto.pack(padx=12, pady=(0, 8))

    # -------------------------
    # Vendedor / Operario (AUTOCOMPLETE desde SQL)
    # -------------------------
    tk.Label(top, text="Vendedor / Operario").pack(anchor="w", padx=12)

    cargar_vendedores_db()

    if DF_VENDEDORES_CACHE is None or DF_VENDEDORES_CACHE.empty:
        messagebox.showwarning(
            "Vendedores",
            "No se pudieron cargar los vendedores desde la base de datos."
        )
        top.destroy()
        return

    lista_vendedores = sorted(
        DF_VENDEDORES_CACHE["NomVen"]
        .dropna()
        .str.strip()
        .str.upper()
        .tolist()
    )

    frame_auto, var_vendedor = crear_autocomplete(
        top,
        lista_vendedores
    )
    frame_auto.pack(padx=12, pady=(0, 12), anchor="w")

    # -------------------------
    # Botones
    # -------------------------
    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=10)

    def confirmar():
        desc = entry_desc.get().strip()
        monto_txt = entry_monto.get().strip()
        vendedor = var_vendedor.get().strip().upper()

        # -------------------------
        # Validaciones
        # -------------------------
        if not desc or not monto_txt or not vendedor:
            messagebox.showwarning(
                "Datos incompletos",
                "Debe completar todos los campos."
            )
            return

        if vendedor not in lista_vendedores:
            messagebox.showwarning(
                "Vendedor inválido",
                "Seleccione un vendedor válido de la lista."
            )
            return

        try:
            monto = float(monto_txt.replace(",", "."))
        except ValueError:
            messagebox.showwarning(
                "Monto inválido",
                "Ingrese un monto numérico válido."
            )
            return

        # -------------------------
        # Guardar anotación
        # -------------------------
        agregar_anotacion(
            nro_transaccion,   # ← número REAL buscado
            vendedor,
            monto,
            desc
        )

        # -------------------------
        # Refrescar TODA la UI
        # (sin destruir widgets primero)
        # -------------------------
        refrescar_calculo_principal()  # ← TU handler global
        refrescar_anotaciones_ui()  # <- Refresca anotaciones
        actualizar_labels_anotaciones()  # ventana / columna anotaciones

        # -------------------------
        # Feedback
        # -------------------------
        messagebox.showinfo(
            "Anotación cargada",
            "La anotación quedó registrada en memoria."
        )

        # -------------------------
        # Cerrar ventana
        # -------------------------
        top.destroy()


    tk.Button(
        frame_btn,
        text="Confirmar",
        bg="#2980b9",
        fg="white",
        font=("Arial", 9, "bold"),
        width=15,
        command=confirmar
    ).pack(side="left", padx=5)

    tk.Button(
        frame_btn,
        text="Cancelar",
        width=15,
        command=top.destroy
    ).pack(side="left", padx=5)


def accion_buscar_transaccion(root):
    if not DATOS_DETALLE_QR:
        messagebox.showwarning(
            "Buscar transacción",
            "Primero debe consultar los datos SQL."
        )
        return

    nro = simpledialog.askstring(
        "Buscar transacción",
        "Ingrese número de transacción:",
        parent=root
    )

    if not nro:
        return

    estado, dato1, dato2 = buscar_transaccion_qr(nro)

    # ✅ QR OK
    if estado == "OK":
        vendedor = dato1
        global TRANSACCION_QR_BUSCADA
        TRANSACCION_QR_BUSCADA = str(nro)
        ver_detalle_qr(vendedor, root)
        return

    # ⚠️ QR SIN COMPROBANTE
    if estado == "SIN_COMPROBANTE":
        fila_qr = dato1
        ventana_asignar_qr_sin_comprobante(root, nro, fila_qr)
        return

    # ❌ NO ENCONTRADA
    ventana_anotacion_visual(nro, root)

def buscar_transaccion_qr(id_transaccion):
    # 🔒 Si ya fue asignado manualmente, excluir del circuito QR
    print("BUSCANDO TRANSACCION:", id_transaccion)
    print("ANOTACIONES:")
    for a in ANOTACIONES_TMP:
        print(a["transaccion"], a["estado"])

    for a in ANOTACIONES_TMP:
        if (
            str(a.get("transaccion")) == id_transaccion
            and a.get("estado") == ESTADO_QR_ASIGNADO_MANUAL
        ):
            return "NO_ENCONTRADA", None, None


    if not id_transaccion:
        return "NO_ENCONTRADA", None, None

    id_transaccion = str(id_transaccion).strip()

    for vendedor, df in DATOS_DETALLE_QR.items():

        if df is None or df.empty:
            continue

        if "ID_TRANSACCION" not in df.columns:
            continue

        coincidencia = df[
            df["ID_TRANSACCION"].astype(str) == id_transaccion
        ]

        if coincidencia.empty:
            continue

        fila = coincidencia.iloc[0]

        comprobante = str(fila.get("NCO", "")).strip()

        vendedor_qr = fila.get("VENDEDOR") or fila.get("Operario") or vendedor

        if not comprobante or comprobante in ("", "None", "0"):
            return "SIN_COMPROBANTE", fila, None

        return "OK", vendedor_qr, fila




    return "NO_ENCONTRADA", None, None

def formatear_moneda_ui(valor):
    try:
        valor = float(valor)
    except:
        valor = 0.0

    return "$ " + f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def ver_editar_rendiciones(vendedor):
    top = tk.Toplevel()
    top.title(f"Rendiciones - {vendedor}")
    top.geometry("450x300")

    tree = ttk.Treeview(
        top,
        columns=("origen", "tipo", "monto"),
        show="headings"
    )
    tree.heading("origen", text="Origen")
    tree.heading("tipo", text="Tipo")
    tree.heading("monto", text="Monto")


    tree.pack(fill="both", expand=True)

    movimientos = DATOS_RENDICIONES.get(vendedor, {}).get("movimientos", [])

    for i, m in enumerate(movimientos):
        tree.insert(
            "",
            "end",
            iid=str(i),
            values=(
                m.get("origen"),
                m.get("tipo", ""),
                formatear_arg(m.get("monto", 0))
            )
        )

    frame_btn = tk.Frame(top)
    frame_btn.pack(pady=5)

    def editar():
        sel = tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        mov = movimientos[idx]

        nuevo = simpledialog.askfloat(
            "Editar monto",
            "Nuevo monto:",
            initialvalue=mov["monto"],
            parent=top
        )
        if nuevo is None:
            return

        mov["monto"] = nuevo
        actualizar_visual_efectivo(vendedor)
        top.destroy()

    def eliminar():
        sel = tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        movimientos.pop(idx)
        actualizar_visual_efectivo(vendedor)
        top.destroy()

    tk.Button(frame_btn, text="✏️ Editar", command=editar).pack(side="left", padx=5)
    tk.Button(frame_btn, text="🗑 Eliminar", command=eliminar).pack(side="left", padx=5)

def agregar_efectivo_manual(vendedor):
    monto = simpledialog.askfloat(
        "Agregar rendición",
        f"Ingrese monto para {vendedor}:",
        minvalue=-9999999,
        parent=sf_global
    )
    if monto is None:
        return

    DATOS_RENDICIONES.setdefault(vendedor, {}).setdefault(
        "movimientos", []
    ).append({
        "origen": "manual",
        "tipo": "Manual",
        "monto": monto
    })

    actualizar_visual_efectivo(vendedor)

def abrir_lapiz_efectivo(vendedor):
    top = tk.Toplevel()
    top.title(f"Efectivo - {vendedor}")
    top.geometry("700x400")

    cols = ("origen", "tipo", "ref", "monto")

    tree = ttk.Treeview(top, columns=cols, show="headings")
    for c in cols:
        tree.heading(c, text=c.capitalize())
        tree.column(c, anchor="center")

    tree.pack(expand=True, fill="both", padx=10, pady=10)

    movimientos = DATOS_RENDICIONES.get(vendedor, {}).get("movimientos", [])

    for i, m in enumerate(movimientos):
        tree.insert(
            "",
            "end",
            iid=str(i),
            values=(m["origen"], m.get("tipo",""), m.get("ref",""), formatear_arg(m["monto"]))
        )

def actualizar_visual_anotaciones():
    for vendedor, lbl in LABELS_ANOTACIONES.items():
        total = sum(
            a["monto"]
            for a in ANOTACIONES_TMP
            if a["vendedor"] == vendedor
        )

        if total > 0:
            lbl.config(text=f"-${total:,.2f}")
        else:
            lbl.config(text="$0.00")

def refrescar_efectivos_ui():
    for vendedor in DATOS_RENDICIONES.keys():
        if vendedor in widgets:
            actualizar_visual_efectivo(vendedor)

def actualizar_visual_efectivo(vendedor):
    if vendedor not in widgets:
        return

    movimientos = DATOS_RENDICIONES.get(vendedor, {}).get("movimientos", [])

    total = sum(
        parse_moneda_robusto(m["monto"])
        for m in movimientos
        if m.get("origen") in ("sql", "manual")
    )

    ee = widgets[vendedor]["ef"]
    ee.config(state="normal")
    ee.delete(0, tk.END)
    ee.insert(0, formatear_moneda_ui(total))
    ee.config(state="readonly")

def cargar_rendiciones_sql():
    if DF_RENDICIONES_CACHE is None or DF_RENDICIONES_CACHE.empty:
        return
    # 🔥 LIMPIAR RENDICIONES SQL ANTERIORES
    for v in widgets.keys():
        DATOS_RENDICIONES[v] = {'movimientos': []}

    for _, row in DF_RENDICIONES_CACHE.iterrows():
        nom_sql = normalizar_texto(row['Vendedor'])
        monto = parse_moneda_robusto(row['Monto_Rendido'])

        for vend_ui in widgets.keys():
            if son_nombres_similares(vend_ui, nom_sql):
                DATOS_RENDICIONES.setdefault(vend_ui, {}).setdefault(
                    'movimientos', []
                ).append({
                    'origen': 'sql',
                    'planilla': row['Planilla'],
                    'nro': row['Nro_Rendicion'],
                    'tipo': 'Rendición',
                    'ref': row['Nro_Transaccion'],
                    'monto': monto
                })
                break
    # =========================
    # ACTUALIZAR EFECTIVO EN UI
    # =========================
    for vend in DATOS_RENDICIONES.keys():
       actualizar_visual_efectivo(vend)
    refrescar_efectivos_ui()



def asegurar_header(container):
    global HEADER_CREADO, frame_header

    if HEADER_CREADO:
        return

    frame_header = tk.Frame(container, bg="#2c3e50")
    frame_header.pack(fill="x", padx=10, pady=(5, 5))

    crear_header_tabla(frame_header)

    HEADER_CREADO = True


def crear_header_tabla(parent):
    headers = [
        ("Vendedor", 20),
        ("Venta Excel", 25),
        ("QR SQL", 14),
        ("Prod. Fact", 12),
        ("Percepción", 12),
        ("Cta. Cte.", 14),
        ("Tarjetas", 12),
        ("Efectivo", 12),
        ("Anotaciones",12),
        ("Diferencia", 12),
        ("¿Con quién trabajé?", 22),
        ("", 10)  # botón confirmar
    ]

    for texto, ancho in headers:
        tk.Label(
            parent,
            text=texto,
            width=ancho,
            bg="#2c3e50",
            fg="white",
            font=("Arial", 9, "bold"),
            anchor="center",
            bd=1,
            relief="ridge"
        ).pack(side="left", padx=1)

def refrescar_interfaz():
    cajas = {}

    for vend, w in widgets.items():
        cid = w["caja_id"]
        cajas.setdefault(cid, []).append(vend)

    for cid, vendedores in cajas.items():
        principal = vendedores[0]
        w = widgets[principal]

        nombre_compuesto = " + ".join(vendedores)

        # 🔹 SOLO si existe lbl_nombre
        if "lbl_nombre" in w:
            w["lbl_nombre"].pack_forget()


        # 🔹 el monto SIEMPRE se refresca
        if "lbl_base" in w:
            w["lbl_base"].config(text=f"${w['base']:,.2f}")

def resetear_estado_global():
    global GNC_GENERAL_TOTAL, GNC_COBERTURAS, GNC_TOTAL_COBERTURAS
    global GNC_PARA_CAJA, GNC_EXTRA_POR_RESPONSABLE
    global UNDO_STACK

    global ANOTACIONES_TMP, LABELS_ANOTACIONES
    global DATOS_RENDICIONES, DATOS_DETALLE_QR, DATOS_DETALLE_FACTURACION
    global TRANSACCION_QR_BUSCADA
    global CALCULAR_HANDLER
    global widgets
    global DATOS_DETALLE_REMITOS
    DATOS_DETALLE_REMITOS = {}
    # -------------------------
    # GNC
    # -------------------------
    GNC_GENERAL_TOTAL = 0.0
    GNC_COBERTURAS = []
    GNC_TOTAL_COBERTURAS = 0.0
    GNC_PARA_CAJA = 0.0
    GNC_EXTRA_POR_RESPONSABLE = {}

    # -------------------------
    # ANOTACIONES
    # -------------------------
    ANOTACIONES_TMP = []
    LABELS_ANOTACIONES = {}

    # -------------------------
    # SQL / QR / FACTURACIÓN
    # -------------------------
    DATOS_RENDICIONES = {}
    DATOS_DETALLE_QR = {}
    DATOS_DETALLE_FACTURACION = {}
    TRANSACCION_QR_BUSCADA = None

    # -------------------------
    # UI / CONTROL
    # -------------------------
    widgets = {}
    CALCULAR_HANDLER = None
    UNDO_STACK = []

def restaurar_estado(estado):
    for vend, data in estado.items():
        if vend not in widgets:
            continue

        w = widgets[vend]

        # 🔹 restaurar base lógica
        w["base"] = data["base"]

        # 🔹 restaurar visual
        if "lbl_base" in w:
            w["lbl_base"].config(
                text=f"${w['base']:,.2f}",
                fg="#16a085"
            )

        def set_entry(e, val):
            e.config(state="normal")
            e.delete(0, tk.END)
            e.insert(0, val)

        set_entry(w["qr"], data["qr"])
        set_entry(w["prod"], data["prod"])
        set_entry(w["per"], data["per"])
        set_entry(w["tarj"], data["tarj"])
        set_entry(w["ef"], data["ef"])

        w["fusionado"] = data["fusionado"]
        w["caja_id"] = data["caja_id"]

        if data["fusionado"]:
            w["row"].configure(bg="#dcdcdc")
            for c in w["row"].winfo_children():
                try:
                    c.configure(state="disabled")
                except:
                    pass
        else:
            w["row"].configure(bg="white")
            for c in w["row"].winfo_children():
                try:
                    c.configure(state="normal")
                except:
                    pass

def deshacer_ultima_fusion():
    global btn_deshacer   # ⬅️ PRIMERA LÍNEA, SIN NADA ANTES

    if not UNDO_STACK:
        return

    estado = UNDO_STACK.pop()
    restaurar_estado(estado["widgets"])

    if btn_deshacer is not None:
        btn_deshacer.destroy()
        btn_deshacer = None

    for v in widgets:
        recalcular_por_cajas()


    refrescar_todos_los_combos()
    refrescar_interfaz()

def snapshot_estado():
    estado = {}
    for v, w in widgets.items():
        estado[v] = {
            "base": w["base"],
            "qr": w["qr"].get(),
            "prod": w["prod"].get(),
            "per": w["per"].get(),
            "tarj": w["tarj"].get(),
            "ef": w["ef"].get(),
            "fusionado": w["fusionado"],
            "caja_id": w["caja_id"]
        }
    return estado


def texto_caja(caja_id):
    return " + ".join(sorted(caja_id))


def refrescar_todos_los_combos():
    for vend, w in widgets.items():
        if w.get("fusionado"):
            continue

        combo = w.get("combo")
        if not combo:
            continue

        combo["values"] = [
            x for x, ww in widgets.items()
            if x != vend and not ww.get("fusionado", False)
        ]

def confirmar_trabajo(vendedor_actual, combo_local):
    global btn_deshacer   # ⬅️ SIEMPRE PRIMERA LÍNEA

    vendedor_sel = combo_local.get().strip()

    # =========================
    # VALIDACIONES
    # =========================
    if not vendedor_sel or vendedor_sel not in widgets:
        messagebox.showwarning("Caja", "Seleccione una caja válida.")
        return

    if vendedor_sel == vendedor_actual:
        messagebox.showwarning("Caja", "No puede unirse consigo mismo.")
        return

    w_base = widgets[vendedor_actual]
    w_sum = widgets[vendedor_sel]

    if w_sum.get("fusionado"):
        messagebox.showwarning("Caja", "Esa caja ya está fusionada.")
        return

    # =========================
    # SNAPSHOT PARA DESHACER
    # =========================
    UNDO_STACK.append({
        "widgets": snapshot_estado()
    })

    # =========================
    # 👉 UNIÓN REAL (SUMA TODO)
    # =========================
    unir_cajas_simple(vendedor_actual, vendedor_sel)

    # =========================
    # BOTÓN DESHACER
    # =========================
    if btn_deshacer is not None:
        btn_deshacer.destroy()

    btn_deshacer = tk.Button(
        w_base["row"],
        text="↩ Deshacer",
        bg="#c0392b",
        fg="white",
        font=("Arial", 8, "bold"),
        command=deshacer_ultima_fusion
    )
    btn_deshacer.pack(side="left", padx=5)

    # =========================
    # RECALCULAR TODO
    # =========================
    refrescar_todos_los_combos()
    refrescar_interfaz()



def unir_cajas_real(id_a, id_b):
    caja_a = CAJAS_DATA[id_a]
    caja_b = CAJAS_DATA[id_b]

    nuevo_id = id_a | id_b

    def unir_df(a, b):
        if a is None:
            return b
        if b is None:
            return a
        return pd.concat([a, b], ignore_index=True)

    CAJAS_DATA[nuevo_id] = {
        "vendedores": caja_a["vendedores"] | caja_b["vendedores"],
        "qr": unir_df(caja_a["qr"], caja_b["qr"]),
        "fact": unir_df(caja_a["fact"], caja_b["fact"]),
        "rend": unir_df(caja_a["rend"], caja_b["rend"]),
        "anot": caja_a["anot"] + caja_b["anot"]
    }

    del CAJAS_DATA[id_a]
    del CAJAS_DATA[id_b]

    return nuevo_id



def cargar_responsables_gnc():
    cargar_vendedores_db()  # Trae todo de SQL
    if DF_VENDEDORES_CACHE is None or DF_VENDEDORES_CACHE.empty:
        return []

    # Limpiamos los nombres y los pasamos a mayúsculas
    df = DF_VENDEDORES_CACHE.copy()
    lista_nombres = df["NomVen"].dropna().str.strip().str.upper().unique().tolist()

    return sorted(lista_nombres)

def crear_autocomplete(parent, lista_opciones):
    frame = tk.Frame(parent)

    var = tk.StringVar()
    entry = tk.Entry(frame, textvariable=var, width=30)
    entry.pack()

    listbox = tk.Listbox(frame, height=5)
    listbox.pack_forget()


    def actualizar_lista(*args):
        texto = var.get().upper()
        listbox.delete(0, tk.END)

        if not texto:
            listbox.pack_forget()
            return

        coincidencias = [
            n for n in lista_opciones if texto in n
        ]

        if coincidencias:
            listbox.pack()
            for n in coincidencias:
                listbox.insert(tk.END, n)
        else:
            listbox.pack_forget()

    def seleccionar(event):
        if listbox.curselection():
            seleccionado = listbox.get(listbox.curselection())
            var.set(seleccionado)
            listbox.pack_forget()

    var.trace_add("write", actualizar_lista)
    listbox.bind("<<ListboxSelect>>", seleccionar)

    return frame, var

def cargar_vendedores_db():
    global DF_VENDEDORES_CACHE
    conn = obtener_conexion_sql()
    if not conn:
        return

    query = "SELECT CodVen, NomVen FROM VENDEDORES"
    DF_VENDEDORES_CACHE = pd.read_sql(query, conn)
    conn.close()
def aplicar_coberturas_gnc_a_bases():

    # 0️⃣ reset
    for w in widgets.values():
        w["gnc_asignado"] = 0.0

    for c in GNC_COBERTURAS:
        responsable = normalizar_texto(c.responsable)
        vendedor_cubre = normalizar_texto(c.vendedor_cubre)

        # 🔥 ACÁ VA EL BLOQUE NUEVO
        key_responsable = None
        key_cubre = None

        for vend in widgets.keys():
            if son_nombres_similares(vend, responsable):
                key_responsable = vend
            if son_nombres_similares(vend, vendedor_cubre):
                key_cubre = vend

        monto = c.total()

        if key_responsable:
            widgets[key_responsable]["gnc_asignado"] -= monto
        if key_cubre:
            widgets[key_cubre]["gnc_asignado"] += monto

    # 2️⃣ recalcular base
    for w in widgets.values():
        w["base"] = (
            w["base_excel"]
            + w.get("gnc_base", 0.0)
            + w["gnc_asignado"]
        )
        w["lbl_base"].config(
            text=formatear_moneda_ui(w["base"]),
            fg="#16a085" if w["base"] >= 0 else "#c0392b"
        )


def recalcular_gnc(
    aforadores_gnc_general,
    coberturas_gnc
):
    """
    Calcula:
    - Total GNC general
    - Total coberturas
    - GNC real para caja
    - Extra por responsable
    """

    global GNC_GENERAL_TOTAL
    global GNC_COBERTURAS
    global GNC_TOTAL_COBERTURAS
    global GNC_PARA_CAJA
    global GNC_EXTRA_POR_RESPONSABLE

    # 1️⃣ GNC GENERAL
    GNC_GENERAL_TOTAL = calcular_gnc_general(aforadores_gnc_general)

    # 2️⃣ COBERTURAS
    GNC_COBERTURAS = coberturas_gnc
    GNC_TOTAL_COBERTURAS = sum(c.total() for c in coberturas_gnc)

    # 3️⃣ GNC QUE VA A CAJA
    GNC_PARA_CAJA = GNC_GENERAL_TOTAL - GNC_TOTAL_COBERTURAS

    # 4️⃣ SUMAR A RESPONSABLES
    GNC_EXTRA_POR_RESPONSABLE = {}

    for c in coberturas_gnc:
        GNC_EXTRA_POR_RESPONSABLE.setdefault(c.responsable, 0.0)
        GNC_EXTRA_POR_RESPONSABLE[c.responsable] += c.total()
    # ================= RESPONSABLE GNC =================
    aplicar_coberturas_gnc_a_bases()
    refrescar_calculo_principal()




def normalizar_texto(texto):
    """Limpia el texto, elimina prefijos numerados y lo convierte a mayúsculas."""
    if not texto or str(texto).lower() == 'nan':
        return ""
    # Elimina patrones como "1 - " o "23-"
    return re.sub(r'^\d+\s*-\s*', '', str(texto)).strip().upper()

def parse_moneda_robusto(valor):
    """Convierte un valor de texto o numérico a float, limpiando símbolos como el $."""
    try:
        if valor is None or str(valor).lower() == 'nan' or valor == '':
            return 0.0
        if isinstance(valor, (float, int)):
            return float(valor)

        s = str(valor).strip()
        # 🔥 REPARACIÓN: Quitar signo peso y espacios que rompen la matemática
        s = s.replace("$", "").replace(" ", "")

        # Caso 1.234,56 (formato europeo/argentino con separador de miles y coma decimal)
        if ',' in s and '.' in s:
            # Si la coma está después del punto, es formato argentino
            if s.rfind(',') > s.rfind('.'):
                s = s.replace(".", "").replace(",", ".")
            else: # Formato yanqui
                s = s.replace(",", "")
        # Caso 1234,56 (formato con coma decimal sin puntos)
        elif ',' in s:
            s = s.replace(",", ".")

        return float(s)
    except:
        return 0.0


def formatear_arg(valor):
    """Formatea un número a formato de moneda ARS (1.234,56) e incluye signo."""
    prefix = "+" if valor > 0.001 else ""
    # "{:,.2f}" genera 1,234.56, luego se invierten , y .
    val_str = "{:,.2f}".format(valor).replace(",", "v").replace(".", ",").replace("v", ".")
    return f"{prefix}{val_str}"


def son_nombres_similares(excel, db):
    ex, db = normalizar_texto(excel), normalizar_texto(db)
    if not ex or not db:
        return False

    pal_ex = set(ex.split())
    pal_db = set(db.split())

    # 🔧 FORZAR: si DB es subconjunto del Excel → MATCH
    if pal_db.issubset(pal_ex):
        return True

    # reglas manuales existentes
    if "ROMAN" in db or "ROMAN" in ex:
        if "VELAZQUEZ" in db and "VELAZQUEZ" in ex:
            return True
        return False

    return pal_ex.issubset(pal_db)


    pal_ex, pal_db = set(ex.split()), set(db.split())
    # Similares si un nombre es subconjunto de las palabras del otro.
    return pal_db.issubset(pal_ex) or pal_ex.issubset(pal_db)


# ------------------------------------------------------------
# 3. VENTANA DETALLE (LUPA)
# ------------------------------------------------------------
def ver_detalle_qr(vendedor_excel, root, reasignado=False):
    global DATOS_DETALLE_QR
    global TRANSACCION_QR_BUSCADA

    df_qr = pd.DataFrame()
    nombre_encontrado = "Desconocido"

    # -------------------------
    # Buscar vendedor en cache QR
    # -------------------------
    for db_name, df in DATOS_DETALLE_QR.items():
        if son_nombres_similares(vendedor_excel, db_name):
            df_qr = df.copy()
            nombre_encontrado = db_name
            break

    if df_qr.empty:
        messagebox.showinfo(
            "Sin Datos QR",
            f"No hay transacciones QR para: {vendedor_excel}"
        )
        return

    # -------------------------
    # 🔍 Filtro por transacción puntual
    # -------------------------
    nro_busqueda = TRANSACCION_QR_BUSCADA

    if nro_busqueda:
        df_qr = df_qr[
            df_qr["ID_TRANSACCION"].astype(str) == str(nro_busqueda)
        ]

        if df_qr.empty:
            messagebox.showinfo(
                "Transacción no encontrada",
                "La transacción no se encontró para este vendedor."
            )
            TRANSACCION_QR_BUSCADA = None
            return

    # -------------------------
    # Detectar SIN COMPROBANTE
    # -------------------------
    def texto_comprobante(r):
        nco = r.get("NCO")

        if pd.isna(nco) or str(nco).strip() == "":
            return "SIN COMPROBANTE"

        tip = "" if pd.isna(r.get("TIP")) else str(r.get("TIP")).strip()
        tco = "" if pd.isna(r.get("TCO")) else str(r.get("TCO")).strip()

        return f"{tip} {tco} {str(nco).strip()}".strip()


    df_qr["COMPROBANTE_TXT"] = df_qr.apply(texto_comprobante, axis=1)

    hay_sin_comprobante = (
        df_qr["COMPROBANTE_TXT"]
        .str.upper()
        .eq("SIN COMPROBANTE")
        .any()
    )

    # -------------------------
    # UI
    # -------------------------
    top = tk.Toplevel(root)
    titulo = "Detalle QR (reasignado)" if reasignado else "Detalle QR"
    top.title(f"{titulo}: {vendedor_excel} (SQL: {nombre_encontrado})")
    top.geometry("1350x520")
    top.grab_set()

    cols = (
        "Comprobante",
        "Fecha",
        "Importe",
        "Cashout",
        "Descuento",
        "Total QR",
        "Transacción"
    )

    tree = ttk.Treeview(top, columns=cols, show="headings")

    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, width=180, anchor="center")

    tree.pack(fill="both", expand=True, padx=5, pady=5)

    total_interfaz = 0.0

    for _, r in df_qr.iterrows():
        fecha = pd.to_datetime(r["FEC"]).strftime("%d/%m/%Y %H:%M")

        tree.insert(
            "",
            "end",
            values=(
                r["COMPROBANTE_TXT"],
                fecha,
                formatear_moneda_ui(r["IMPORTE"]),
                formatear_moneda_ui(r["CASHOUT"]),
                formatear_moneda_ui(r["DESC_PROMO_NUM"]),
                formatear_moneda_ui(r["QR_FINAL"]),
                r["ID_TRANSACCION"]
            )
        )

        total_interfaz += float(r["QR_FINAL"])

    # -------------------------
    # TOTAL
    # -------------------------
    bg = "#8e44ad" if reasignado else "#27ae60"

    tk.Label(
        top,
        text=f"TOTAL QR: {formatear_arg(total_interfaz)}",
        font=("Arial", 13, "bold"),
        bg=bg,
        fg="white"
    ).pack(side="bottom", fill="x")

    # -------------------------
    # 🟡 BOTÓN ASIGNAR (solo si SIN COMPROBANTE y una sola fila)
    # -------------------------
    if hay_sin_comprobante and len(df_qr) == 1:

        def asignar_qr_a_vendedor():
            fila_qr = df_qr.iloc[0]
            nro = str(fila_qr["ID_TRANSACCION"])

            top.destroy()

            ventana_asignar_qr_sin_comprobante(
                root,
                nro,
                fila_qr
            )

        tk.Button(
            top,
            text="🟡 ASIGNAR QR A VENDEDOR",
            bg="#f1c40f",
            fg="black",
            font=("Arial", 10, "bold"),
            height=2,
            command=asignar_qr_a_vendedor
        ).pack(fill="x", padx=10, pady=6)

    TRANSACCION_QR_BUSCADA = None


def ver_detalle_remitos(vendedor_excel, root):
    """Muestra una ventana con el detalle de remitos (Cuentas Corrientes) para un vendedor."""
    global DATOS_DETALLE_REMITOS
    df_vendedor = pd.DataFrame()
    nombre_encontrado = "Desconocido"

    for db_name, df in DATOS_DETALLE_REMITOS.items():
        if son_nombres_similares(vendedor_excel, db_name):
            df_vendedor = df
            nombre_encontrado = db_name
            break

    if df_vendedor.empty:
        messagebox.showinfo("Sin Datos", f"No hay remitos asignados a: {vendedor_excel}")
        return

    top = tk.Toplevel(root)
    top.title(f"Detalle Cuentas Corrientes: {vendedor_excel} (En SQL: {nombre_encontrado})")
    top.geometry("1000x450") # La hice un poco más ancha para que entre el cliente

    # 1. Nuevas columnas
    cols = ("Fecha", "Comprobante", "Cliente", "Vendedor", "Total")
    tree = ttk.Treeview(top, columns=cols, show='headings')

    # 2. Configurar anchos y alineaciones para que se vea prolijo
    anchos = {"Fecha": 120, "Comprobante": 120, "Cliente": 250, "Vendedor": 150, "Total": 100}
    for c in cols:
        tree.heading(c, text=c)
        # El cliente lo alineamos a la izquierda ("w"), los demás centrados
        alineacion = "w" if c == "Cliente" else "center"
        tree.column(c, width=anchos.get(c, 100), anchor=alineacion)

    scroll = ttk.Scrollbar(top, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scroll.set)

    tree.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    total = 0.0
    for _, row in df_vendedor.iterrows():
        monto = row['Total_Remito']
        total += monto

        # Formatear fecha y hora
        fecha_str = pd.to_datetime(row['Fecha']).strftime("%d/%m/%Y %H:%M") if pd.notna(row['Fecha']) else ""

        # Si el cliente está vacío, le ponemos Consumidor Final
        cliente = str(row['Cliente']).strip() if pd.notna(row['Cliente']) and str(row['Cliente']).strip() != "" else "CONSUMIDOR FINAL"

        vals = (fecha_str, row['Comprobante'], cliente, row['Nombre_Vendedor'], f"${monto:,.2f}")
        tree.insert("", "end", values=vals)

    tk.Label(top, text=f"TOTAL CUENTAS CORRIENTES: ${total:,.2f}",
             font=("Arial", 14, "bold"), bg="#3498db", fg="white").pack(side="bottom", fill="x")
def ver_detalle_vendedor(vendedor_excel, root):
    """Muestra una ventana Toplevel con el detalle de facturación SQL para un vendedor."""
    global DATOS_DETALLE_FACTURACION
    df_vendedor = pd.DataFrame()
    nombre_encontrado = "Desconocido"

    # Buscar el vendedor en los datos cargados por SQL
    for db_name, df in DATOS_DETALLE_FACTURACION.items():
        if son_nombres_similares(vendedor_excel, db_name):
            df_vendedor = df
            nombre_encontrado = db_name
            break

    if df_vendedor.empty:
        messagebox.showinfo("Sin Datos", f"No hay datos SQL asignados a: {vendedor_excel}")
        return

    top = tk.Toplevel(root)
    top.title(f"Detalle: {vendedor_excel} (En SQL figura como: {nombre_encontrado})")
    top.geometry("1100x500")

    cols = ("Fecha", "Tipo", "Numero", "Ref_CPA", "Vendedor_Origen", "Producto", "Total_Neto")
    tree = ttk.Treeview(top, columns=cols, show='headings')

    # Configuración de columnas
    for c in cols:
        tree.heading(c, text=c)
        if c == "Producto":
            tree.column(c, width=250)
        elif c == "Total_Neto":
            tree.column(c, width=100, anchor="e")
        else:
            tree.column(c, width=90)

    # Scrollbar
    scroll = ttk.Scrollbar(top, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scroll.set)

    # Empaquetado
    tree.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    total = 0.0
    for _, row in df_vendedor.iterrows():
        monto = row['Total_Neto']
        total += monto

        vals = (
            row['Fecha'], row['Tipo'], row['Numero'], row['Ref_CPA'],
            row['Vendedor_Nombre'],
            row['Producto'], f"${monto:,.2f}"
        )
        # Etiqueta para colorear NC en rojo
        tag = 'nc' if (monto < 0 or "NC" in str(row['Tipo'])) else 'ft'
        tree.insert("", "end", values=vals, tags=(tag,))

    # Configuración de color para Tags
    tree.tag_configure('nc', foreground='red')
    tree.tag_configure('ft', foreground='black')

    # Etiqueta de total final
    bg = "#27ae60" if total >= 0 else "#c0392b"
    tk.Label(top, text=f"TOTAL NETO FINAL: ${total:,.2f}",
             font=("Arial", 14, "bold"), bg=bg, fg="white").pack(side="bottom", fill="x")


# ------------------------------------------------------------
# 4. INTERFAZ PRINCIPAL
# ------------------------------------------------------------
def mostrar_planilla(df=None, root=None):
    global CONSULTAR_SQL_HANDLER
    global widgets
    global CALCULAR_HANDLER
    global EJECUTAR_SQL_HANDLER
    MODO_PISAR_EXCEL = False


    frame_cobertura_inner = None

    widgets = {}
    def asegurar_fila(vendedor, importe_inicial=0.0, color="#ffffff"):
        vendedor = normalizar_texto(vendedor)
        asegurar_header(frame_header_container)



        # 🟢 YA EXISTE
        if vendedor in widgets:
            w = widgets[vendedor]

            if MODO_PISAR_EXCEL:
                # 🔄 PISAR Excel
                w["base_excel"] = importe_inicial
                w["base_original"] = importe_inicial
            else:
                # ➕ SUMAR (otros orígenes: GNC, coberturas, etc.)
                w["base_excel"] += importe_inicial
                w["base_original"] += importe_inicial

            # Recalcular base final
            w["base"] = w["base_excel"] + w.get("gnc_asignado", 0.0)

            w["lbl_base"].config(text=formatear_arg(w["base"]))
            w["lbl_monto"].config(text=formatear_arg(w["base_excel"]))

            aplicar_coberturas_gnc_a_bases()
            refrescar_interfaz()
            return w



        # 🔵 NO EXISTE → crear fila nueva
        row = tk.Frame(sf, bg=color, bd=1, relief="groove")
        row.pack(fill="x", pady=2)

        tk.Label(row, text=vendedor, width=20, anchor="w", bg=color).pack(side="left", padx=2)

        lbl_nombre = tk.Label(
            row,
            text=vendedor,
            anchor="w",
            width=28,
            fg="#16a085",
            font=("Arial", 9)
        )
        lbl_nombre.pack(side="left", padx=(5, 10))

        lbl_monto = tk.Label(
            row,
            text=f"${importe_inicial:,.2f}",   # ✅ CORREGIDO
            anchor="e",
            width=12,
            fg="#16a085",
            font=("Arial", 9)
        )
        lbl_monto.pack(side="left", padx=(0, 10))

        lbl_base = tk.Label(
            row,
            text=f"${importe_inicial:,.2f}",
            width=12,
            bg=color,
            font=("Arial", 9, "bold"),
            fg="#16a085"
        )
        lbl_base.pack(side="left", padx=2)

        # QR SQL (Entrada + Botón Detalle)
        qr_f = tk.Frame(row, bg=color)
        qr_f.pack(side="left", padx=2)

        eq = tk.Entry(qr_f, width=11, justify="center", state="disabled")
        eq.pack(side="left")

        tk.Button(
            qr_f,
            text="🔍",
            font=("Arial", 7),
            bg="#ecf0f1",
            command=lambda x=vendedor: abrir_detalle_qr(x, vent)

        ).pack(side="left")


        pf_f = tk.Frame(row, bg=color)
        pf_f.pack(side="left", padx=2)

        ep = tk.Entry(pf_f, width=11, justify="center")
        ep.pack(side="left")

        tk.Button(
            pf_f,
            text="🔍",
            font=("Arial", 7),
            command=lambda x=vendedor: ver_detalle_vendedor(x, vent)
        ).pack(side="left")

        eper = tk.Entry(row, width=12, justify="center")
        eper.pack(side="left", padx=2)
        eper.insert(0, "0")
        eper.config(bg="#f4f6f7")


        # --------- NUEVO: CUENTAS CORRIENTES (Remitos) ---------
        cta_f = tk.Frame(row, bg=color)
        cta_f.pack(side="left", padx=2)

        ecta = tk.Entry(cta_f, width=11, justify="center", state="readonly")
        ecta.pack(side="left")

        tk.Button(
            cta_f, text="🔍", font=("Arial", 7), bg="#ecf0f1",
            command=lambda x=vendedor: ver_detalle_remitos(x, vent)
        ).pack(side="left")
        # -------------------------------------------------------

        et = tk.Entry(row, width=12, justify="center")
        et.pack(side="left", padx=2)
        ef_frame = tk.Frame(row, bg=color)
        ef_frame.pack(side="left", padx=2)
        # -------------------------
        # ANOTACIONES
        # -------------------------
        lbl_anotaciones = tk.Label(
            row,
            text="$0.00",
            width=12,
            anchor="e",
            font=("Arial", 9, "bold"),
            fg="#c0392b",
            bg=color
        )
        lbl_anotaciones.pack(side="left", padx=2)

        vend_norm = normalizar_texto(vendedor)
        LABELS_ANOTACIONES[vend_norm] = lbl_anotaciones


        ee = tk.Entry(
            ef_frame,
            width=11,
            justify="right",
            font=("Segoe UI", 9),
            state="readonly",
            readonlybackground="white",
            relief="solid"
)

        ee.pack(side="left")
        btn_mas = tk.Button(
            ef_frame,
            text="+",
            width=2,
            height=1,
            bg=COLOR_MAS,
            fg="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat",
            command=lambda v=vendedor: agregar_efectivo_manual(v)
        )
        btn_mas.pack(side="left")

        btn_edit = tk.Button(
            ef_frame,
            text="✏",
            width=2,
            height=1,
            bg=COLOR_EDIT,
            fg="black",
            font=("Segoe UI", 9),
            relief="flat",
            command=lambda v=vendedor: ver_editar_rendiciones(v)
        )
        btn_edit.pack(side="left")



        ld = tk.Label(row, text="0,00", width=12, bg="#f8f9fa", font=("Arial", 9, "bold"))
        ld.pack(side="left", padx=2)

        frame_trab = tk.Frame(row, bg=color)
        frame_trab.pack(side="left", padx=5)

        combo = ttk.Combobox(frame_trab, state="disabled", width=18)
        combo.pack(side="left")

        tk.Button(
            frame_trab,
            text="Confirmar",
            font=("Arial", 8, "bold"),
            bg="#34495e",
            fg="white",
            command=lambda vend=vendedor, cb=combo: confirmar_trabajo(vend, cb)
        ).pack(side="left", padx=3)

        widgets[vendedor] = {
            "qr": eq,
            "prod": ep,
            "per": eper,
            "cta_cte": ecta,
            "tarj": et,
            "ef": ee,
            "diff": ld,
            # 🔹 BASES SEPARADAS
            "base_excel": importe_inicial,   # ventas Excel reales
            "gnc_asignado": 0.0,              # gnc neto (post coberturas)

            "base": importe_inicial,
            "base_original": importe_inicial,
            "lbl_base": lbl_base,
            "row": row,
            "fusionado": False,
            "gnc_aplicado": False,
            "combo": combo,
            "caja_id": frozenset([vendedor]),
            "lbl_nombre": lbl_nombre,
            "lbl_monto": lbl_monto
        }
        cid = widgets[vendedor]["caja_id"]

        CAJAS_DATA[cid] = {
            "vendedores": {vendedor},
            "qr": DATOS_DETALLE_QR.get(vendedor),              # DataFrame
            "fact": DATOS_DETALLE_FACTURACION.get(vendedor),  # DataFrame
            "rend": DATOS_RENDICIONES.get(vendedor),           # DataFrame o lista
            "anot": [a for a in ANOTACIONES_TMP if a["vendedor"] == vendedor]
        }



        refrescar_todos_los_combos()
        refrescar_interfaz()
        return widgets[vendedor]

    """Crea la interfaz principal con la tabla de auditoría para cada vendedor."""
    vent = root
    vent.title("Auditoría Contable - Con Corrección de NC")
    vent.geometry("1360x900")
    vent.state("zoomed")
    vent.configure(bg="#f4f7f6")


    # Contenedor y Canvas para permitir scroll
    # CONTENEDOR RAÍZ
    container = tk.Frame(vent, bg="#f4f7f6")
    container.pack(fill="both", expand=True)

    # ZONA SUPERIOR (GNC + Coberturas)
    frame_top = tk.Frame(container, bg="#f4f7f6")
    frame_top.pack(fill="x", side="top")

    # ZONA CENTRAL (tabla con scroll)
    frame_center = tk.Frame(container, bg="#f4f7f6")
    frame_center.pack(fill="both", expand=True)

    # ZONA INFERIOR (BOTONES FIJOS)
    frame_bottom = tk.Frame(container, bg="#f4f7f6")
    frame_bottom.pack(fill="x", side="bottom")


    canvas = tk.Canvas(container, bg="#f4f7f6")
    scroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    # ===== CONTENEDORES FIJOS =====
    frame_aforadores = tk.Frame(frame_top, bg="#f4f7f6")
    frame_aforadores.pack(fill="x", padx=10, pady=(10, 5))

    # 🔹 CONTENEDOR HORIZONTAL PARA GNC + COBERTURA
    frame_aforadores_h = tk.Frame(frame_aforadores, bg="#f4f7f6")
    frame_aforadores_h.pack(fill="x")
    frame_aforadores_h.pack_propagate(False)
    frame_aforadores_h.config(height=380)

    frame_anotaciones = tk.Frame(frame_bottom, bg=frame_bottom["bg"])
    frame_anotaciones.pack(side="left", padx=6)

    btn_anotaciones = tk.Button(
    frame_anotaciones,
    text="📝 Anotaciones",
    bg="#8e44ad",
    fg="white",
    font=("Arial", 9, "bold"),
    command=lambda: abrir_anotaciones(vent)
    )
    btn_anotaciones.pack(side="left", padx=4)

    btn_actualizar = tk.Button(
    frame_anotaciones,
    text="🔄 Actualizar",
    bg="#16a085",
    fg="white",
    font=("Arial", 9, "bold"),
    command=actualizar_completo
    )
    btn_actualizar.pack(side="left", padx=4)



    # =========================================================
    # PANEL COBERTURA GNC (DERECHA DEL GNC)
    # =========================================================
    frame_cobertura = tk.LabelFrame(
        frame_aforadores_h,   # 👈 MISMO CONTENEDOR HORIZONTAL
        text="Cobertura GNC",
        font=("Arial", 10, "bold"),
        bg="#ecf0f1",
        padx=10,
        pady=10
    )

    frame_cobertura_visible = False
    frame_cobertura_inner = tk.Frame(frame_cobertura, bg="#ecf0f1")
    frame_cobertura_inner.pack(fill="x", anchor="w")

    # CONTENEDOR HORIZONTAL: cobertura | baños | botón
    frame_cobertura_row = tk.Frame(frame_cobertura_inner, bg="#ecf0f1")
    frame_cobertura_row.pack(anchor="w", fill="x")

        # CONTENEDOR HORIZONTAL DE AFORADORES (cobertura + baños)
  #  frame_aforadores_cobertura = tk.Frame(frame_cobertura_inner, bg="#ecf0f1")
  #  frame_aforadores_cobertura.pack(anchor="w", fill="x")


    frame_header_container = tk.Frame(frame_center, bg="#f4f7f6")
    frame_header_container.pack(fill="x", padx=10, pady=(0, 5))

    frame_scroll = tk.Frame(frame_center, bg="#f4f7f6")
    frame_scroll.pack(fill="both", expand=True)


    # ===== SCROLL SOLO PARA FILAS =====
    canvas = tk.Canvas(frame_scroll, bg="#f4f7f6")
    scroll = ttk.Scrollbar(frame_scroll, orient="vertical", command=canvas.yview)

    sf = tk.Frame(canvas, bg="#f4f7f6")  # SOLO FILAS
    sf_global = sf

    sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=sf, anchor="nw")
    canvas.configure(yscrollcommand=scroll.set)

    canvas.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")


    widgets = {}

    if df is not None:
        vendedores = sorted(df['Vendedor'].dropna().unique())
    else:
        vendedores = []




    # --- Filas de Vendedores ---
    for v in vendedores:

        row = tk.Frame(sf, bg="white", bd=1, relief="groove")
        row.pack(fill="x", pady=2)
        tot_ex = df[df['Vendedor'] == v]['Importe'].sum()

        # Vendedor (Excel)
        tk.Label(row, text=v, width=20, anchor="w", bg="white").pack(side="left", padx=2)
        # Total Venta (Excel)
        lbl_base = tk.Label(
            row,
            text=f"${tot_ex:,.2f}",
            width=12,
            bg="white",
            font=("Arial",9,"bold"),
            fg="#16a085"
        )
        lbl_base.pack(side="left", padx=2)


        # QR SQL (Entrada + Botón Detalle)
        qr_f = tk.Frame(row, bg="white")
        qr_f.pack(side="left", padx=2)

        eq = tk.Entry(qr_f, width=11, justify="center", state="disabled")
        eq.pack(side="left")

        tk.Button(
            qr_f,
            text="🔍",
            font=("Arial", 7),
            bg="#ecf0f1",
            command=lambda x=v: ver_detalle_qr(x, vent)
        ).pack(side="left")


        # Productos Facturados (Entrada + Botón Detalle)
        pf_f = tk.Frame(row, bg="white")
        pf_f.pack(side="left", padx=2)
        ep = tk.Entry(pf_f, width=11, justify="center",state="readonly")
        ep.pack(side="left")

        # Percepción (solo DB)
        eper = tk.Entry(row, width=12, justify="center")
        eper.pack(side="left", padx=2)
        eper.insert(0, "0")
        eper.config(bg="#f4f6f7")



        # Se usa lambda con argumento por defecto para capturar el valor correcto de 'v'
        tk.Button(pf_f, text="🔍", font=("Arial", 7), bg="#ecf0f1",
                 command=lambda x=v: ver_detalle_vendedor(x, vent)).pack(side="left")

        # Tarjetas (Entrada)
        et = tk.Entry(row, width=12, justify="center")
        et.pack(side="left", padx=2)

        # Efectivo (solo DB)

        var_ef = tk.StringVar(value="0,00")

        ee = tk.Entry(
            row,
            textvariable=var_ef,
            width=12,
            justify="center",
            state="readonly",
            readonlybackground="white"
        )

        ee.pack(side="left", padx=2)
        # EFECTIVO
        # -------------------------
        # ANOTACIONES
        # -------------------------
        lbl_anotaciones = tk.Label(
            row,
            text=formatear_moneda_ui(total_anotaciones_por_vendedor(v)),
            width=12,
            anchor="e",
            font=("Arial", 9, "bold"),
            fg="#c0392b",
            bg="white"
        )
        lbl_anotaciones.pack(side="left", padx=2)

        LABELS_ANOTACIONES[v] = lbl_anotaciones


        # Diferencia (Label de Resultado)

        ld = tk.Label(row, text="0,00", width=12, bg="#f8f9fa", font=("Arial", 9, "bold"))
        ld.pack(side="left", padx=2)




        # --- ¿Con quién trabajé? (por fila) ---
        frame_trab = tk.Frame(row, bg="white")
        frame_trab.pack(side="left", padx=5)

        combo = ttk.Combobox(
            frame_trab,
            values=[],
            width=18,
            state="readonly"
        )
        combo.pack(side="left")

        tk.Button(
            frame_trab,
            text="Confirmar",
            font=("Arial", 8, "bold"),
            bg="#34495e",
            fg="white",
            command=lambda vend=v, cb=combo: confirmar_trabajo(vend, cb)
        ).pack(side="left", padx=3)

        widgets[v] = {"qr": eq, "prod": ep, "tarj": et,"per": eper, "ef": ee, "diff": ld, "base": tot_ex, "base_original": tot_ex,"lbl_base":lbl_base, "gnc_aplicado": False,"caja_id": frozenset([v])
,"row":row, "fusionado":False,"combo": combo}
    refrescar_todos_los_combos()
    # ======= TOTAL CAJA CONJUNTA =======
    lbl_total_caja_conjunta = tk.Label(
        frame_bottom,
        text="Total Caja Conjunta: $0,00",
        font=("Arial", 12, "bold"),
        bg="#f4f7f6",
        fg="#d35400"
    )
    lbl_total_caja_conjunta.pack(side="left", padx=10, pady=5)

    def actualizar_total_caja_conjunta():
        total = TOTAL_CAJA_CONJUNTA_VALOR
        color = "#27ae60" if total >= 0 else "#c0392b"

        lbl_total_caja_conjunta.config(
            text=f"Total Caja Conjunta: ${total:,.2f}",
            fg=color
        )

    global TOTAL_CAJA_CONJUNTA_FN
    TOTAL_CAJA_CONJUNTA_FN = actualizar_total_caja_conjunta
    # =========================================================
    # PANEL GNC
    # =========================================================
    frame_gnc = tk.LabelFrame(
    frame_aforadores_h,   # 🔥 CAMBIO CLAVE
    text="GNC - Aforadores",
    font=("Arial", 10, "bold"),
    bg="#ecf0f1",
    padx=10,
    pady=10
    )
    frame_gnc.pack(
        side="left",
        fill="y",
        padx=(0, 10),
        pady=10
    )


    def mostrar_cobertura_gnc():
        nonlocal frame_cobertura_visible

        if not frame_cobertura_visible:
            refrescar_combo_cobertura()
            frame_cobertura.pack(
                side="left",
                fill="y",
                padx=(10, 0),
                pady=10
            )
            frame_cobertura_visible = True
        else:
            frame_cobertura.pack_forget()
            frame_cobertura_visible = False


    tk.Button(
    frame_gnc,   # 👈 se queda al lado del GNC
    text="Agregar cobertura GNC +",
    bg="#8e44ad",
    fg="white",
    font=("Arial", 9, "bold"),
    command=mostrar_cobertura_gnc
).pack(anchor="w", pady=(5, 10))




    tk.Label(
        frame_cobertura,
        text="Responsable:",
        bg="#ecf0f1",
        font=("Arial", 9, "bold")
    ).pack(anchor="w")

    combo_cobertura = ttk.Combobox(
        frame_cobertura,
        state="readonly",
        width=25
    )
    combo_cobertura.pack(anchor="w", pady=5)
    def obtener_responsables_desde_widgets():
        """
        Responsables posibles para Cobertura GNC,
        tomados de la planilla (widgets).
        """
        responsables = set()

        for nombre, w in widgets.items():
            if not w.get("fusionado", False):
                responsables.add(nombre)

        return sorted(responsables)


    def refrescar_combo_cobertura():
        valores = obtener_responsables_desde_widgets()
        combo_cobertura["values"] = valores

        actual = combo_cobertura.get()
        if actual not in valores:
            combo_cobertura.set("")


    tk.Label(
        frame_cobertura_inner,
        text="Aforadores Cobertura",
        bg="#ecf0f1",
        font=("Arial", 9, "bold")
    ).pack(anchor="w", pady=(5, 2))

    aforadores_cobertura_entries = []


    for i in range(4):
        f = tk.Frame(frame_cobertura_row, bg="#ecf0f1")


        f.pack(anchor="w", pady=2)

        tk.Label(f, text=f"Aforador {i+1}", width=10, bg="#ecf0f1").pack(side="left")

        e_ini = tk.Entry(f, width=8, justify="center")
        e_ini.pack(side="left", padx=3)
        e_ini.insert(0, "0")

        e_fin = tk.Entry(f, width=8, justify="center")
        e_fin.pack(side="left", padx=3)
        e_fin.insert(0, "0")

        lbl_val = tk.Label(f, text="$0,00", width=10, bg="#ecf0f1")
        lbl_val.pack(side="left", padx=3)

        aforadores_cobertura_entries.append((e_ini, e_fin, lbl_val))
    var_banio_cobertura = tk.BooleanVar()
    # CONTENEDOR INTERNO: aforadores cobertura + baños (MISMA ALTURA)



    frame_banio = tk.Frame(frame_cobertura_row, bg="#ecf0f1")


    frame_banio.pack_forget()

    def toggle_banio():
        if var_banio_cobertura.get():
            frame_banio.pack(side="left", padx=(15, 0), anchor="n")
        else:
            frame_banio.pack_forget()



    tk.Checkbutton(
        frame_cobertura,
        text="Baños / doble cobertura",
        variable=var_banio_cobertura,
        command=toggle_banio,
        bg="#ecf0f1"
    ).pack(anchor="w", pady=5)

    aforadores_banio_entries = []

    tk.Label(
        frame_banio,
        text="Aforadores Baños",
        bg="#ecf0f1",
        font=("Arial", 9, "bold")
    ).pack(anchor="w")

    for i in range(4):
        f = tk.Frame(frame_banio, bg="#ecf0f1")
        f.pack(anchor="w", pady=2)

        tk.Label(f, text=f"Aforador {i+1}", width=10, bg="#ecf0f1").pack(side="left")

        e_ini = tk.Entry(f, width=8, justify="center")
        e_ini.pack(side="left", padx=3)
        e_ini.insert(0, "0")

        e_fin = tk.Entry(f, width=8, justify="center")
        e_fin.pack(side="left", padx=3)
        e_fin.insert(0, "0")

        lbl_val = tk.Label(f, text="$0,00", width=10, bg="#ecf0f1")
        lbl_val.pack(side="left", padx=3)

        aforadores_banio_entries.append((e_ini, e_fin, lbl_val))



    def confirmar_cobertura_gnc():
        global GNC_COBERTURAS

        vendedor_cubre = normalizar_texto(combo_cobertura.get())
        if not vendedor_cubre:
            messagebox.showwarning("Cobertura", "Seleccione quién cubre")
            return

        responsable_gnc = normalizar_texto(var_responsable_gnc.get())
        if not responsable_gnc:
            messagebox.showwarning("Cobertura", "No hay responsable GNC definido")
            return

        precio = parse_moneda_robusto(entry_precio_gnc.get())
        GNC_COBERTURAS.clear()

        afs = []
        for ini, fin, lbl in aforadores_cobertura_entries:
            a = AforadorGNC(
                parse_moneda_robusto(ini.get()),
                parse_moneda_robusto(fin.get()),
                precio
            )
            lbl.config(text=f"${a.total():,.2f}")
            afs.append(a)

        if afs:
            c = CoberturaGNC(
                nombre="Cobertura GNC",
                responsable=responsable_gnc,
                aforadores=afs
            )
            c.vendedor_cubre = vendedor_cubre   # ✅ atributo agregado
            GNC_COBERTURAS.append(c)

        # --- Baños ---
        if var_banio_cobertura.get():
            afs_banio = []
            for ini, fin, lbl in aforadores_banio_entries:
                a = AforadorGNC(
                    parse_moneda_robusto(ini.get()),
                    parse_moneda_robusto(fin.get()),
                    precio
                )
                lbl.config(text=f"${a.total():,.2f}")
                afs_banio.append(a)

            if afs_banio:
                c = CoberturaGNC(
                    nombre="Baños",
                    responsable=responsable_gnc,
                    aforadores=afs_banio
                )
                c.vendedor_cubre = vendedor_cubre   # ✅ idem
                GNC_COBERTURAS.append(c)

        aplicar_coberturas_gnc_a_bases()
        calcular()

        messagebox.showinfo(
            "Cobertura GNC",
            "Cobertura aplicada correctamente."
        )



    btn_confirmar_cobertura = tk.Button(
        frame_cobertura_row,
        text="Confirmar\nCobertura",
        bg="#27ae60",
        fg="white",
        font=("Arial", 9, "bold"),
        width=8,
        height=3,
        command=confirmar_cobertura_gnc
    )
    btn_confirmar_cobertura.pack(side="left", padx=(10, 0), anchor="n")


        # --- Precio GNC ---
    precio_frame = tk.Frame(frame_gnc, bg="#ecf0f1")
    precio_frame.pack(fill="x", pady=(0, 10))
        # ================= RESPONSABLE GNC =================
        # ================= RESPONSABLE GNC =================
    def agregar_responsable_gnc():
        global RESPONSABLE_GNC_ACTUAL

        responsable = normalizar_texto(var_responsable_gnc.get())

        if not responsable:
            messagebox.showwarning("GNC", "Ingrese un responsable.")
            return
        RESPONSABLE_GNC_ACTUAL = responsable
        if GNC_PARA_CAJA <= 0:
            messagebox.showwarning(
                "GNC",
                "Primero debe calcular el GNC (aforadores).\n"
                "El monto para caja es $0."
            )
            return

        total_gnc = GNC_PARA_CAJA

        # ======================================================
        # 1️⃣ SI YA EXISTE → SUMAR GNC
        # ======================================================
        for vend_existente, w in widgets.items():
            if son_nombres_similares(vend_existente, responsable):

                if w.get("gnc_aplicado", False):
                    messagebox.showwarning(
                        "GNC",
                        f"El GNC ya fue aplicado a:\n{vend_existente}"
                    )
                    return

                w["gnc_base"] = total_gnc



                aplicar_coberturas_gnc_a_bases()



                w["gnc_aplicado"] = True

                w["lbl_base"].config(
                text=formatear_moneda_ui(w["base"]),
                fg="#d35400"
            )



                refrescar_todos_los_combos()

                messagebox.showinfo(
                    "GNC",
                    f"GNC (${total_gnc:,.2f}) agregado a:\n{vend_existente}"
                )
                return


        # ======================================================
        # 2️⃣ NO EXISTE → CREAR FILA USANDO PIPELINE NORMAL
        # ======================================================
        w = asegurar_fila(
            responsable,
            importe_inicial=total_gnc,
            color="#fff9e6"
        )

        w["gnc_aplicado"] = True

        # 🔹 asegurar estructura de rendiciones
        DATOS_RENDICIONES.setdefault(responsable, {"movimientos": []})

        # 🔹 inyectar GNC SQL (si existe)
        inyectar_gnc_a_responsable(responsable)

        # 🔹 refrescar efectivo (habilita + y ✏️)
        refrescar_efectivos_ui()

        # 🔹 recalcular todo
        refrescar_calculo_principal()

        messagebox.showinfo(
            "GNC",
            f"{responsable} no existía.\n"
            f"Se agregó con GNC (${total_gnc:,.2f})."
        )
    def actualizar_gnc_en_responsable():
        if not RESPONSABLE_GNC_ACTUAL:
            return

        total_gnc = GNC_PARA_CAJA

        for vend, w in widgets.items():
            if son_nombres_similares(vend, RESPONSABLE_GNC_ACTUAL):
                if w.get("gnc_aplicado", False):
                    w["gnc_base"] = total_gnc
                    aplicar_coberturas_gnc_a_bases()

                    w["lbl_base"].config(
                        text=formatear_moneda_ui(w["base"]),
                        fg="#d35400"
                    )

                    refrescar_calculo_principal()
                return

    responsables_gnc = cargar_responsables_gnc()
    frame_resp = tk.Frame(frame_gnc, bg="#ecf0f1")
    frame_resp.pack(anchor="w", pady=5)


    frame_auto, var_responsable_gnc = crear_autocomplete(
        frame_resp,
        responsables_gnc
    )
    frame_auto.pack(side="left")

    tk.Button(
        frame_resp,
        text="Agregar GNC ➕",
        bg="#d35400",
        fg="white",
        font=("Arial", 9, "bold"),
        command=agregar_responsable_gnc
    ).pack(side="left", padx=5)


    tk.Label(
        frame_gnc,
        text="Responsable GNC:",
        font=("Arial", 10, "bold"),
        bg="#ecf0f1"
    ).pack(anchor="w", pady=(10, 2))

    # ================= COBERTURA BAÑO =================


    def calcular_gnc_ui():
        """
        Lee los aforadores GNC desde la UI,
        recalcula totales y actualiza labels.
        """
        aforadores = []

        precio = parse_moneda_robusto(entry_precio_gnc.get())

        for e_ini, e_fin, lbl_valor in aforadores_entries:
            ini = parse_moneda_robusto(e_ini.get())
            fin = parse_moneda_robusto(e_fin.get())

            a = AforadorGNC(
                inicial=ini,
                final=fin,
                precio=precio
            )

            aforadores.append(a)

            # Valor individual
            lbl_valor.config(text=f"${a.total():,.2f}")

        recalcular_gnc(
            aforadores_gnc_general=aforadores,
            coberturas_gnc=[]
        )

        lbl_gnc_total.config(
            text=f"Total GNC: ${GNC_GENERAL_TOTAL:,.2f}"
        )

        lbl_gnc_caja.config(
            text=f"GNC para Caja: ${GNC_PARA_CAJA:,.2f}"
        )
        actualizar_gnc_en_responsable()

    tk.Button(
        frame_gnc,
        text="Calcular GNC ⛽",
        bg="#27ae60",
        fg="white",
        font=("Arial", 10, "bold"),
        command=calcular_gnc_ui
    ).pack(pady=10,  anchor="w")


    # --- Funciones de Botones ---



    tk.Label(
        precio_frame,
        text="Precio GNC ($):",
        width=15,
        anchor="w",
        bg="#ecf0f1",
        font=("Arial", 9, "bold")
    ).pack(side="left")

    entry_precio_gnc = tk.Entry(precio_frame, width=10, justify="center")
    entry_precio_gnc.pack(side="left")
    entry_precio_gnc.insert(0, str(PRECIO_GNC_DEFAULT))

    aforadores_entries = []

    for i in range(4):
        f = tk.Frame(frame_gnc, bg="#ecf0f1")
        f.pack(fill="x", pady=2)

        tk.Label(f, text=f"Aforador {i+1}", width=12, bg="#ecf0f1").pack(side="left")

        e_ini = tk.Entry(f, width=10, justify="center")
        e_ini.pack(side="left", padx=5)
        e_ini.insert(0, "0")

        e_fin = tk.Entry(f, width=10, justify="center")
        e_fin.pack(side="left", padx=5)
        e_fin.insert(0, "0")

        lbl_valor = tk.Label(
            f,
            text="$0,00",
            width=12,
            anchor="e",
            bg="#ecf0f1",
            font=("Arial", 9, "bold")
        )
        lbl_valor.pack(side="left", padx=5)

        aforadores_entries.append((e_ini, e_fin, lbl_valor))


    # Resultados
    lbl_gnc_total = tk.Label(frame_gnc, text="Total GNC: $0,00",
                             font=("Arial", 10, "bold"), bg="#ecf0f1")
    lbl_gnc_total.pack(anchor="w", pady=5)

    lbl_gnc_caja = tk.Label(frame_gnc, text="GNC para Caja: $0,00",
                            font=("Arial", 10, "bold"), bg="#ecf0f1")
    lbl_gnc_caja.pack(anchor="w")




    def consultar_sql_completo():
        """Pide el rango de planillas y ejecuta las consultas SQL de QR y Facturación/NC."""
        global PLANILLA_DESDE_SQL, PLANILLA_HASTA_SQL
        global DATOS_DETALLE_FACTURACION

        root.attributes("-topmost", True)

        d = simpledialog.askinteger(
            "Rango",
            "Planilla DESDE:",
            parent=root
        )
        if d is None:
            root.attributes("-topmost", False)
            return

        h = simpledialog.askinteger(
            "Rango",
            "Planilla HASTA:",
            initialvalue=d,
            parent=root
        )
        if h is None:
            root.attributes("-topmost", False)
            return

        root.attributes("-topmost", False)


        messagebox.showinfo("Proceso completado",
                            "QR y total facturado fueron procesados correctamente.")
        PLANILLA_DESDE_SQL = d
        PLANILLA_HASTA_SQL = h
        ejecutar_consulta_sql_con_planillas(d, h)

    def ejecutar_consulta_sql_con_planillas(d, h):
        global DATOS_DETALLE_FACTURACION
        conn = obtener_conexion_sql()
        if not conn:
            return

        global DATOS_DETALLE_FACTURACION
        DATOS_DETALLE_FACTURACION = {}

        try:
            # Consulta QR (Sin cambios)
          #  sql_qr = """
           # WITH Piezas AS (
            #    SELECT COALESCE(v1.NOMVEN, v2.NOMVEN, 'ID: ' + CAST(af.OPE AS VARCHAR)) AS Operario,
             #   TRY_CAST(REPLACE(REPLACE(A.IMPORTE, '.', ''), ',', '.') AS DECIMAL(18,2))/100 AS Neto,
              #  TRY_CAST(REPLACE(REPLACE(ISNULL(A.CASHOUT, '0'), '.', ''), ',', '.') AS DECIMAL(18,2))/100 AS Cash
               # FROM A_MERCADOPAGO A
                #JOIN AMAEFACT_EXT C ON A.CAR = C.ID_COMPV2
                #JOIN amaefact af ON af.SUC = C.SUC AND af.NCO = C.NCO AND af.TIP = C.TIP AND af.TCO = C.TCO
               # LEFT JOIN VENDEDORES v1 ON af.VEN = v1.CODVEN
               # LEFT JOIN VENDEDORES v2 ON af.OPE = v2.CODVEN
               # WHERE A.PLA BETWEEN ? AND ? AND A.Est IN ('1', '11', '12')
           # )
           # SELECT Operario, SUM(Neto) + SUM(Cash) as Total FROM Piezas GROUP BY Operario
           # """

            # Consulta FACTURACIÓN + NC (CORREGIDA)
            sql_qr = """
            ;WITH QR_BASE AS (
    SELECT
        A.PLA AS Planilla,

        COALESCE(v2.NOMVEN, v1.NOMVEN, 'SIN OPERARIO') AS Operario,

        A.NTRANS AS ID_TRANSACCION,

        CAST(REPLACE(A.IMPORTE, ',', '.') AS DECIMAL(18,2)) AS IMPORTE,
        CAST(REPLACE(ISNULL(A.CASHOUT, '0'), ',', '.') AS DECIMAL(18,2)) AS CASHOUT,

        A.RED_DES AS DESC_PROMO,
        CAST(REPLACE(ISNULL(A.RED_TOT,'0'), ',', '.') AS DECIMAL(18,2)) AS DESCUENTO_TOTAL,

        A.FEC,
        A.ORI,
        A.LUG,
        A.EST AS EST_MP,

        A.CAR AS ID_COMPV2,
        af.TIP,
        af.TCO,
        af.NCO
    FROM A_MERCADOPAGO A
    LEFT JOIN AMAEFACT_EXT C
        ON A.CAR = C.ID_COMPV2
    LEFT JOIN AMAEFACT af
        ON af.SUC = C.SUC
       AND af.NCO = C.NCO
       AND af.TIP = C.TIP
       AND af.TCO = C.TCO
    LEFT JOIN VENDEDORES v1 ON af.VEN = v1.CODVEN
    LEFT JOIN VENDEDORES v2 ON af.OPE = v2.CODVEN
    WHERE
        A.PLA BETWEEN ? AND ?
        OR A.LUG = 5
        AND A.LUG = 1
        AND A.EST = 1
),

-- 🔹 Depuración de duplicados: nos quedamos solo con el “real”
QR_DEPURADO AS (
    SELECT *,
           ROW_NUMBER() OVER (
               PARTITION BY ID_COMPV2
               ORDER BY NCO DESC
           ) AS rn
    FROM QR_BASE
),

-- 🔹 Clasificación final de QR
QR_CLASIFICADO AS (
    SELECT *,
        CASE
            WHEN ID_COMPV2 IS NULL THEN 'QR SIN COMPROBANTE'
            WHEN NCO IS NULL THEN 'QR COMPROBANTE INEXISTENTE'
            ELSE 'QR OK'
        END AS ESTADO_QR,
        ROUND(IMPORTE + CASHOUT - DESCUENTO_TOTAL, 2) AS QR_NETO_FINAL
    FROM QR_DEPURADO
    WHERE rn = 1 -- elimina duplicados/fantasmas
)

-- =========================
-- DETALLE POR VENDEDOR
-- =========================
SELECT
    Planilla,
    Operario,
    ESTADO_QR,
    ID_TRANSACCION,
    IMPORTE,
    CASHOUT,
    DESC_PROMO,
    DESCUENTO_TOTAL,
    QR_NETO_FINAL,
    FEC,
    TIP,
    TCO,
    NCO
FROM QR_CLASIFICADO
ORDER BY
    Operario,
    FEC;




                    """


            sql_facturas = """
                                SELECT
            f.FEC AS Fecha,
            f.TCO AS Tipo,
            f.NCO AS Numero,
            f.CPA AS Ref_CPA,
            COALESCE(v_origen.NomVen, v_actual.NomVen, 'DESCONOCIDO') AS Vendedor_Nombre,
            art.DetArt AS Producto,
            CAST(d.CAN AS DECIMAL(18,2)) AS Cantidad,
            CASE
                WHEN f.TCO LIKE '%NC%' THEN
                    (CAST(d.CAN AS DECIMAL(18,2)) * -1) * CAST(art.PreVen AS DECIMAL(18,2))
                ELSE
                    CAST(d.CAN AS DECIMAL(18,2)) * CAST(art.PreVen AS DECIMAL(18,2))
            END AS Total_Neto
        FROM AMAEFACT f
        INNER JOIN AMOVSTOC d
            ON f.SUC = d.PVE AND f.NCO = d.NCO
            AND f.TIP = d.TIP AND f.TCO = d.TCO
        INNER JOIN ARTICULOS art ON d.ART = art.CodArt
        LEFT JOIN AMAEFACT f_origen
            ON f.SUC = f_origen.SUC
            AND f.TCO LIKE '%NC%'
            AND f_origen.TCO NOT LIKE '%NC%'
            AND TRY_CAST(f.CPA AS BIGINT) = f_origen.NCO
        LEFT JOIN VENDEDORES v_origen ON f_origen.OPE = v_origen.CodVen
        LEFT JOIN VENDEDORES v_actual ON f.OPE = v_actual.CodVen
        WHERE f.PLA BETWEEN ? AND ?-- Filtro de planilla
            AND f.ANU = ''
            AND art.DetArt <> 'GNC' -- <--- ESTO QUITA EL PRODUCTO GNC DE LA LISTA
            AND (
                -- CASO 1: Si la planilla tiene artículos de PLAYA, mostrar solo PLAYA
                (
                    EXISTS (
                        SELECT 1 FROM AMOVSTOC d2
                        INNER JOIN ARTICULOS a2 ON d2.ART = a2.CodArt
                        WHERE d2.PVE = f.SUC AND d2.NCO = f.NCO AND d2.TCO = f.TCO
                        AND (a2.codSec = 0 AND a2.CodRub = 6 OR a2.DetArt LIKE '%PELOTA%' OR (a2.CodRub = 10 AND d2.ART IN (2,4)))
                    )
                    AND (art.codSec = 0 AND art.CodRub = 6 OR art.DetArt LIKE '%PELOTA%' OR (art.CodRub = 10 AND d.ART IN (2,4)))
                )
                OR
                -- CASO 2: Si NO tiene nada de playa, entonces intentar mostrar el resto del Sector 1 Rubro 1
                (
                    NOT EXISTS (
                        SELECT 1 FROM AMOVSTOC d3
                        INNER JOIN ARTICULOS a3 ON d3.ART = a3.CodArt
                        INNER JOIN AMAEFACT f3 ON f3.SUC = d3.PVE AND f3.NCO = d3.NCO AND f3.TCO = d3.TCO
                        WHERE f3.PLA = f.PLA
                        AND (a3.codSec = 0 AND a3.CodRub = 6 OR a3.DetArt LIKE '%PELOTA%' OR (a3.CodRub = 10 AND d3.ART IN (2,4)))
                    )
                    AND (art.codSec = 1 AND art.CodRub = 1)
                )
            )
        ORDER BY f.FEC, f.NCO;
        """
            sql_percepcion = """
            SELECT
                f.TCO,
                CAST(f.SUC AS VARCHAR(5)) + '-' + CAST(f.NCO AS VARCHAR(20)) AS Comprobante,
                f.TOT AS Total_Factura,
                f.PER AS Percepcion,
                v.NomVen AS Nombre_Vendedor
            FROM AMAEFACT f
            LEFT JOIN VENDEDORES v
                ON f.OPE = v.CodVen
            WHERE f.PLA BETWEEN ? AND ?  -- Corregido para usar el rango
            AND f.TCO LIKE 'F%'
            AND f.TIP = 'A'
            AND f.ANU = ''
            ORDER BY f.SUC, f.NCO
            """
            sql_efectivo = """
            SELECT
                r.PLA AS Planilla,
                r.NUM AS Nro_Mov,
                r.OPE AS Operario,
                v.NomVen AS Vendedor,
                r.FEC AS Fecha,
                r.EFE AS Efectivo,
                r.LUG AS Sector
            FROM ATURRPA r
            LEFT JOIN VENDEDORES v
                ON r.OPE = v.CodVen
            WHERE r.PLA BETWEEN ? AND ?
            AND r.LUG = 1
            ORDER BY v.NomVen, r.FEC
            """
            sql_remito = """
            SELECT
                f.FEC AS Fecha,
                f.TCO AS Tipo,
                CAST(f.SUC AS VARCHAR(5)) + '-' + CAST(f.NCO AS VARCHAR(20)) AS Comprobante,
                f.NOM AS Cliente,
                f.TOT AS Total_Remito,
                v.NomVen AS Nombre_Vendedor
            FROM AMAEFACT f
            LEFT JOIN VENDEDORES v ON f.OPE = v.CodVen
            WHERE f.PLA BETWEEN ? AND ?
                AND f.TCO = 'RE'
                AND f.ANU = ''
            ORDER BY f.FEC, f.SUC, f.NCO
            """


            df_efectivo = pd.read_sql(
                sql_efectivo,
                conn,
                params=[d, h]
            )


            global DF_GNC_SQL_CACHE
            DF_GNC_SQL_CACHE = df_efectivo.copy()



            print("--- CONSULTANDO SQL ---")
            df_qr = pd.read_sql(sql_qr, conn, params=(d, h))

            # =========================
            # NORMALIZAR NUMÉRICOS
            # =========================
            df_qr["IMPORTE"] = pd.to_numeric(df_qr["IMPORTE"], errors="coerce").fillna(0.0)
            df_qr["CASHOUT"] = pd.to_numeric(df_qr["CASHOUT"], errors="coerce").fillna(0.0)

            # =========================
            # NORMALIZAR PROMO (🔥 CLAVE)
            # =========================


            df_qr["DESC_PROMO_NUM"] = df_qr["DESC_PROMO"].apply(normalizar_desc_promo)
            df_qr = (
            df_qr
            .groupby("ID_TRANSACCION", as_index=False)
            .agg({
                "Operario": "first",
                "IMPORTE": "first",
                "CASHOUT": "first",
                "DESC_PROMO": "first",
                "DESC_PROMO_NUM": "sum",   # 🔥 SUMA REAL
                "FEC": "first",
                "ESTADO_QR": "first",
                "TIP": "first",
                "TCO": "first",
                "NCO": "first",
            })
        )
            # =========================
            # DEBUG PROMOS
            # =========================



            df_dbg = df_qr[df_qr["DESC_PROMO_NUM"] > 0][
                    ["Operario", "IMPORTE", "CASHOUT", "DESC_PROMO_NUM"]
                ]


            print(
                df_qr[df_qr["DESC_PROMO_NUM"] > 0][
                    ["ID_TRANSACCION", "IMPORTE", "DESC_PROMO_NUM"]
                ].head(10)
            )



            print("=== DEBUG PROMOS ===")
            print(df_dbg.head(20))
            print("MAX PROMO:", df_dbg["DESC_PROMO_NUM"].max())
            print("MIN PROMO:", df_dbg["DESC_PROMO_NUM"].min())

            # =========================
            # CÁLCULO FINAL QR
            # =========================
            df_qr["QR_FINAL"] = (
                df_qr["IMPORTE"]
                + df_qr["CASHOUT"]
                - df_qr["DESC_PROMO_NUM"]
            ).round(2)

            total_qr = round(df_qr["QR_FINAL"].sum(), 2)

            # =========================
            # GUARDAR DETALLE QR
            # =========================
            DF_RENDICIONES_CACHE = pd.read_sql(sql_efectivo, conn, params=(d, h))
            # =========================
            # LIMPIAR EFECTIVO SQL PREVIO (IDEMPOTENCIA)
            # =========================
            for vend in DATOS_RENDICIONES:
                if 'movimientos' in DATOS_RENDICIONES[vend]:
                    DATOS_RENDICIONES[vend]['movimientos'] = [
                        m for m in DATOS_RENDICIONES[vend]['movimientos']
                        if m.get('origen') != 'sql'
        ]

            # =========================
            # CARGAR EFECTIVO SQL A MOVIMIENTOS
            # =========================
            for _, row in DF_RENDICIONES_CACHE.iterrows():
                vendedor_sql = str(row['Vendedor']).strip().upper()
                monto = parse_moneda_robusto(row['Efectivo'])

                for vend_ui in widgets.keys():
                    if son_nombres_similares(vend_ui, vendedor_sql):
                        DATOS_RENDICIONES.setdefault(vend_ui, {}).setdefault('movimientos', []).append({
                            'origen': 'sql',
                            'planilla': row['Planilla'],
                            'nro': row['Nro_Mov'],
                            'tipo': 'Efectivo',
                            'ref': f"Mov {row['Nro_Mov']} - {row['Fecha']:%d/%m %H:%M}",
                            'monto': float(monto)
                        })
                        break

            DATOS_DETALLE_QR.clear()
            refrescar_efectivos_ui()

            DATOS_DETALLE_QR.clear()

            for operario in df_qr['Operario'].dropna().unique():

               # 1️⃣ Crear df del operario
                df_qr_vendedor = df_qr[df_qr['Operario'] == operario].copy()

                # 2️⃣ Mantener QR OK y QR SIN COMPROBANTE
                df_qr_vendedor = df_qr_vendedor[
                    df_qr_vendedor["ESTADO_QR"].isin([
                        "QR OK",
                        "QR SIN COMPROBANTE",
                        "QR COMPROBANTE INEXISTENTE"
                    ])
                ]

                # 3️⃣ Solo guardar si quedó algo
                if not df_qr_vendedor.empty:

                    # 🔒 evitar duplicar la misma transacción QR
                    df_qr_vendedor = df_qr_vendedor.drop_duplicates(
                        subset=["ID_TRANSACCION"],
                        keep="first"
                    )

                    DATOS_DETALLE_QR[operario] = df_qr_vendedor


            global DESCUENTOS_QR_POR_VENDEDOR
            DESCUENTOS_QR_POR_VENDEDOR = {}
            for operario, df in DATOS_DETALLE_QR.items():
               DESCUENTOS_QR_POR_VENDEDOR[operario] = df["DESC_PROMO_NUM"].sum()
            # =========================
            # CREAR ÍNDICE DE QR (para anotaciones)
            # =========================
            INDICE_QR = {}

            for operario, df in DATOS_DETALLE_QR.items():
                for _, row in df.iterrows():

                    # indexar por ID_TRANSACCION
                    if pd.notna(row.get("ID_TRANSACCION")):
                        INDICE_QR[str(row["ID_TRANSACCION"])] = {
                            "operario_original": operario,
                            "qr_row": row
                        }

                    # indexar por NCO (si existe)
                    if pd.notna(row.get("NCO")):
                        INDICE_QR[str(row["NCO"])] = {
                            "operario_original": operario,
                            "qr_row": row
                        }


            # =========================
            # INYECTAR QR EN CAJAS
            # =========================
            for vendedor_sql, df in DATOS_DETALLE_QR.items():
                total_qr = round(df["QR_FINAL"].sum(), 2)




                for vend_ui, w in widgets.items():
                    if son_nombres_similares(vend_ui, vendedor_sql):

                        # 🔹 valor lógico para el cálculo
                        w["qr_db"] = total_qr

                        # 🔹 reflejo visual
                        eq = w["qr"]
                        eq.config(state="normal")
                        eq.delete(0, tk.END)
                        eq.insert(0, f"{total_qr:.2f}")
                        eq.config(state="disabled")

                        break

            df_fact = pd.read_sql(sql_facturas, conn, params=(d, h))

            # La consulta de percepciones debe usar el rango, no solo el 'desde' (d)
            df_per = pd.read_sql(sql_percepcion, conn, params=(d, h))

            # Creación de df_per_sum (esto es vital y ahora está ANTES de usarse)
            df_per_sum = (
                df_per
                .groupby('Nombre_Vendedor', dropna=False)['Percepcion']
                .sum()
                .reset_index()
            )

            # EJECUTAR CONSULTA REMITOS
            df_remito = pd.read_sql(sql_remito, conn, params=(d, h))
            DATOS_DETALLE_REMITOS.clear()

            # Guardar df por vendedor
            for v_sql in df_remito['Nombre_Vendedor'].dropna().unique():
                DATOS_DETALLE_REMITOS[v_sql] = df_remito[df_remito['Nombre_Vendedor'] == v_sql]

            # Sumarizar e inyectar en la UI
            df_remito_sum = df_remito.groupby('Nombre_Vendedor')['Total_Remito'].sum().reset_index()

            for _, r in df_remito_sum.iterrows():
                nom_sql = r['Nombre_Vendedor']
                monto = r['Total_Remito']

                for v_ex, w in widgets.items():
                    if son_nombres_similares(v_ex, nom_sql):
                        w['cta_cte'].config(state="normal")
                        w['cta_cte'].delete(0, tk.END)
                        w['cta_cte'].insert(0, f"{monto:.2f}")
                        w['cta_cte'].config(state="readonly")
            conn.close()

            # ---------------------------------------------------------
            # PURGA GLOBAL DE NC + FACTURAS ANULADAS (POR NUMERO)
            # ---------------------------------------------------------
            # ... (Toda la lógica de purga se mantiene igual)

            # 1. Detectar Notas de Crédito
            df_nc = df_fact[df_fact['Tipo'].str.contains('NC', na=False)].copy()

            # 2. Obtener todos los números de comprobantes anulados desde Ref_CPA
            numeros_anulados = set()

            for _, row in df_nc.iterrows():
                try:
                    ref = row['Ref_CPA']
                    if pd.notna(ref):
                        ref_num = int(str(ref).strip())
                        numeros_anulados.add(ref_num)
                except:
                    pass

            print(f"Comprobantes anulados detectados: {numeros_anulados}")

            # 3. Eliminar
            indices_a_borrar = []

            for idx, row in df_fact.iterrows():
                es_nc = 'NC' in str(row['Tipo'])
                try:
                    numero = int(str(row['Numero']).strip())
                except:
                    numero = -1

                if es_nc or numero in numeros_anulados:
                    indices_a_borrar.append(idx)

            df_fact = df_fact.drop(indices_a_borrar)
            DATOS_PRODUCTOS_FACTURADOS = df_fact.copy()
            print(f"Filas luego de purga NC: {len(df_fact)}")

            # ... (Debugging)
            print(f"Total filas encontradas: {len(df_fact)}")

            ncs_check = df_fact[df_fact['Tipo'].str.contains('NC', na=False)]
            if not ncs_check.empty:
                print(f"\n¡SE ENCONTRARON {len(ncs_check)} NOTAS DE CRÉDITO!")
                print("Mira a quién se asignaron:")
                print(ncs_check[['Fecha', 'Numero', 'Ref_CPA', 'Vendedor_Nombre', 'Total_Neto']])
            else:
                print("\nNO se encontraron Notas de Crédito en este rango de planillas.")

            # 1. Llenar QR
           # c_qr = 0
           # for _, r in df_qr.iterrows():
           #     for v_ex, w in widgets.items():
            #        if son_nombres_similares(v_ex, r['Operario']):
             #           w['qr'].delete(0, tk.END)
              #          w['qr'].insert(0, f"{r['Total']:.2f}")
               #         c_qr += 1



            # 2. Llenar Percepción (USANDO df_per_sum, que ya existe)
            for _, r in df_per_sum.iterrows():
                nom_sql = r['Nombre_Vendedor']
                monto = parse_moneda_robusto(r['Percepcion'])

                for v_ex, w in widgets.items():
                    if son_nombres_similares(v_ex, nom_sql):
                        w['per'].delete(0, tk.END)
                        w['per'].insert(0, f"{monto:.2f}")

            # 3. Llenar Facturación (Productos)
            for v_sql in df_fact['Vendedor_Nombre'].unique():
                DATOS_DETALLE_FACTURACION[v_sql] = df_fact[df_fact['Vendedor_Nombre'] == v_sql]

            df_sumas = df_fact.groupby('Vendedor_Nombre')['Total_Neto'].sum().reset_index()

            c_fact = 0
            for _, r in df_sumas.iterrows():
                nom_sql = r['Vendedor_Nombre']
                monto = r['Total_Neto']

                for v_ex, w in widgets.items():
                    if son_nombres_similares(v_ex, nom_sql):
                        w['prod'].delete(0, tk.END)
                        w['prod'].insert(0, f"{monto:.2f}")
                        c_fact += 1


        except Exception as e:
            messagebox.showerror("Error SQL", str(e))
            print("ERROR FATAL:", e)

    CONSULTAR_SQL_HANDLER = consultar_sql_completo
    EJECUTAR_SQL_HANDLER = ejecutar_consulta_sql_con_planillas
    def calcular():
        # =========================
        # AGRUPAR POR CAJA
        # =========================
        cajas = {}

        for v, w in widgets.items():
            cid = w["caja_id"]
            cajas.setdefault(cid, []).append((v, w))

        # =========================
        # CALCULAR POR CAJA
        # =========================
        global TOTAL_CAJA_CONJUNTA_VALOR
        TOTAL_CAJA_CONJUNTA_VALOR = 0.0

        for cid, lista in cajas.items():

            entrada_bruta = 0.0
            salida_real = 0.0
            ajustes = 0.0

            fila_principal = None

            for v, w in lista:
                if not w.get("fusionado"):
                    fila_principal = w

                # =========================
                # ENTRADA ESPERADA
                # =========================
                base = w["base"]  # 👈 YA incluye Excel + GNC + Coberturas
                prod = parse_moneda_robusto(w["prod"].get())
                per  = parse_moneda_robusto(w["per"].get())

                entrada_bruta += base + prod + per

                # =========================
                # SALIDA REAL
                # =========================
                qr = w.get("qr_db", 0.0)

                movs = DATOS_RENDICIONES.get(v, {}).get("movimientos", [])
                efectivo = sum(
                    parse_moneda_robusto(m["monto"])
                    for m in movs
                    if m.get("origen") in ("sql", "manual")
                )

                tarj = parse_moneda_robusto(w["tarj"].get())
                cta_cte = parse_moneda_robusto(w.get("cta_cte").get()) # <--- LEER VALOR

                # SUMAR cta_cte A LA SALIDA
                salida_real += qr + efectivo + tarj + cta_cte

                # =========================
                # AJUSTES
                # =========================
                ajustes += total_anotaciones_por_vendedor(v)

                promo_qr = 0.0
                for vend_sql, total_promo in DESCUENTOS_QR_POR_VENDEDOR.items():
                    if son_nombres_similares(v, vend_sql):
                        promo_qr = total_promo
                        break

                ajustes += promo_qr

            # =========================
            # RESULTADO FINAL
            # =========================
            entrada_real = entrada_bruta - ajustes
            diferencia = salida_real - entrada_real
            # 🔥 SOLO vendedores principales (no fusionados)
            if not w.get("fusionado", False):
                TOTAL_CAJA_CONJUNTA_VALOR += diferencia

            DIFERENCIAS_VENDEDORES[v] = diferencia
            color = "#27ae60" if diferencia >= 0 else "#e74c3c"

            if fila_principal:
                fila_principal["diff"].config(
                    text=formatear_arg(diferencia),
                    fg=color
                )

       # Limpiar fusionadas
            for v, w in lista:
                if w is not fila_principal:
                    w["diff"].config(text="", fg="black")

        # =======================================================
        # 🔹 SUMAR QR "SIN OPERARIO" A CAJA CONJUNTA
        # =======================================================
        qr_sin_operario = 0.0
        # Buscamos exactamente la clave 'SIN OPERARIO' que definimos en el SQL
        if "SIN OPERARIO" in DATOS_DETALLE_QR and DATOS_DETALLE_QR["SIN OPERARIO"] is not None:
            qr_sin_operario = DATOS_DETALLE_QR["SIN OPERARIO"]["QR_FINAL"].sum()

        # Le sumamos ese dinero digital huérfano al total de la estación
        TOTAL_CAJA_CONJUNTA_VALOR += qr_sin_operario

        # =========================
        # 🔹 TOTAL CAJA CONJUNTA (UNA SOLA VEZ)
        # =========================
        if TOTAL_CAJA_CONJUNTA_FN:
            TOTAL_CAJA_CONJUNTA_FN()

        global ACTUALIZAR_DIFERENCIAS_FN
        ACTUALIZAR_DIFERENCIAS_FN = actualizar_diferencias_ui
    CALCULAR_HANDLER = calcular
    # --- Botones de Control ---
    #et.pack(side="left", padx=20)
    bf = tk.Frame(frame_bottom, bg="#f4f7f6")
    bf.pack(pady=10, anchor="center")



    def abrir_excel_ui():
        nonlocal MODO_PISAR_EXCEL
        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls")]
        )
        if not path:
            return

        h, p = encontrar_encabezado_y_columnas(path)
        if h == -1:
            return

        df = pd.read_excel(path, header=h)
        df.rename(columns={df.columns[p]: 'Importe'}, inplace=True)

        if 'Tipo de Venta' in df.columns:
            df = df[~df['Tipo de Venta'].astype(str)
                    .str.contains('YER', case=False, na=False)]

        df['Importe'] = df['Importe'].apply(parse_moneda_robusto)
        MODO_PISAR_EXCEL = True
        inyectar_excel_en_planilla(df)
        MODO_PISAR_EXCEL = False
    # Alias provisorios de vendedores del Excel → nombre real en el sistema
    ALIAS_VENDEDORES_EXCEL = {}
    def inyectar_excel_en_planilla(df):
        vendedores_excel = df.groupby("Vendedor")["Importe"].sum()

        for vend, importe in vendedores_excel.items():
            if not vend:
                continue

            vend_norm = str(vend).strip().upper()

            # 🔁 Aplicar alias si corresponde
            vend_final = ALIAS_VENDEDORES_EXCEL.get(vend_norm, vend_norm)

            asegurar_fila(vend_final, importe)

        # 🔥 CLAVE: reaplicar anotaciones ya cargadas
        if ANOTACIONES_TMP:
            actualizar_labels_anotaciones()

        refrescar_todos_los_combos()
        refrescar_combo_cobertura()


    def reiniciar_interfaz_ui():
        if not messagebox.askyesno(
            "Reiniciar",
            "¿Borrar todo y comenzar de nuevo?"
        ):
            return

        resetear_estado_global()

        # 🔥 limpiar UI sin destruir ventana
        for child in vent.winfo_children():
            child.destroy()

        mostrar_planilla(None, vent)

    tk.Button(
        bf,
        text="📂 Abrir Excel",
        bg="#2980b9",
        fg="white",
        font=("Arial", 10, "bold"),
        command=abrir_excel_ui
    ).pack(side="left", padx=20)

    tk.Button(
        bf,
        text="🔄 Reiniciar interfaz",
        bg="#c0392b",
        fg="white",
        font=("Arial", 10, "bold"),
        command=reiniciar_interfaz_ui
    ).pack(side="left", padx=20)




    tk.Button(bf, text="Consultar SQL 🔄", bg="#8e44ad", fg="white",
              font=("Arial", 10,"bold"), command=consultar_sql_completo).pack(side="left", padx=20)

    tk.Button(bf, text="Calcular 🧮", bg="#3498db", fg="white",
              font=("Arial", 10,"bold"), command=calcular).pack(side="left", padx=20)

    tk.Button(
    bf,
    text="Guardar Cierre 💾",
    bg="#2980b9",
    fg="white",
    font=("Arial", 10, "bold"),
    relief="raised",
    command=lambda:abrir_ventana_guardado(root)
).pack(side="left", padx=20)




# ------------------------------------------------------------
# 5. MAIN
# ------------------------------------------------------------
def encontrar_encabezado_y_columnas(path):
    raw = pd.read_excel(path, header=None)

    for i, row in raw.iterrows():
        fila = row.astype(str).str.lower()

        # Detectar fila de encabezados
        if fila.str.contains('vendedor').any() and fila.str.contains('fecha').any():

            # Buscar columna de importe
            posibles = fila[fila.str.contains('importe|total|pesos', regex=True)]

            if not posibles.empty:
                return i, posibles.index[0]
            else:
                messagebox.showerror(
                    "Error de formato",
                    "Se encontró el encabezado, pero no la columna IMPORTE / TOTAL / PESOS."
                )
                return -1, -1

    messagebox.showerror(
        "Error de formato",
        "No se pudo detectar la fila de encabezados del Excel."
    )
    return -1, -1


def obtener_fecha_internet():
    """Consulta la hora real a los servidores de Google para evitar trampas con el reloj local."""
    try:
        # Hacemos una petición rápida solo para leer el encabezado de la fecha
        req = urllib.request.Request("http://www.google.com", method="HEAD")
        with urllib.request.urlopen(req, timeout=5) as response:
            date_str = response.headers['Date']
            # El formato que devuelve Google es: 'Thu, 19 Feb 2026 16:00:00 GMT'
            fecha_real = datetime.strptime(date_str, "%a, %d %b %Y %H:%M:%S %Z")
            return fecha_real
    except Exception:
        # Si falla (ej: no hay internet), devolvemos None
        return None

def validar_licencia():
    """Verifica si el programa superó la fecha límite de uso."""
    # ----------------------------------------------------
    # 🔥 CONFIGURA TU FECHA DE CADUCIDAD AQUÍ (Año, Mes, Día)
    # ----------------------------------------------------
    fecha_caducidad = datetime(2027, 2, 19)

    fecha_actual = obtener_fecha_internet()

    # ¿Qué pasa si justo se corta el internet en la estación?
    # Usamos la hora local como "salvavidas" temporal para que puedan cerrar la caja igual.
    if not fecha_actual:
        fecha_actual = datetime.now()

    if fecha_actual > fecha_caducidad:
        root = tk.Tk()
        root.withdraw() # Oculta la ventana principal
        messagebox.showerror(
            "Licencia Expirada",
            "El período de uso de este sistema ha finalizado.\n"
            "Por favor, contacte al administrador o desarrollador para renovar la licencia."
        )
        sys.exit() # Cierra el proceso por completo y no deja abrir la app
def main():
    # 1. Ejecutar el control de seguridad antes de cargar nada
    validar_licencia()

    # 2. Cargar el programa normal
    root = tk.Tk()
    root.title("Auditoría Contable")

    print("MAIN ARRANCÓ")

    mostrar_planilla(None, root)

    root.protocol(
        "WM_DELETE_WINDOW",
        lambda: on_cerrar_programa(root)
    )

    root.mainloop()

if __name__ == "__main__":
    main()
